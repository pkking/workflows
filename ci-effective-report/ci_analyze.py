#!/usr/bin/env python3
"""
CI 效率分析报告生成器

从 Turso DB 读取 GitHub Actions 数据，生成 CI 效率分析报告（Excel）。
支持单仓库分析、多仓库对比、时间范围过滤。

用法:
  # 单仓库（默认 vllm-ascend，最近 30 天）
  python ci_analyze.py

  # 指定仓库
  python ci_analyze.py --repo vllm-project/vllm-ascend

  # 多仓库对比
  python ci_analyze.py --repo vllm-project/vllm-ascend --repo vllm-project/vllm

  # 指定时间范围
  python ci_analyze.py --repo vllm-project/vllm-ascend --from 2026-05-01 --to 2026-05-23

  # 仅列出可用仓库
  python ci_analyze.py --list-repos

  # 自定义 step 分类映射
  python ci_analyze.py --step-names my-step-names.json

  # 跳过 Excel 输出（仅打印到终端）
  python ci_analyze.py --no-excel
"""

import argparse
import json
import math
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests

# ─── 配置 ──────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
ENV_FILE = SCRIPT_DIR / ".env"
DEFAULT_STEP_NAMES = SCRIPT_DIR / "step-names.json"
DEFAULT_DAYS = 30


def load_env(path: Path) -> dict:
    env = {}
    if path.exists():
        for line in path.read_text().splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip('"')
    return env


# ─── Turso HTTP 客户端 ─────────────────────────────────────────────────

class TursoClient:
    """Thin wrapper around Turso v2/pipeline HTTP API."""

    def __init__(self, db_url: str, auth_token: str):
        self.url = db_url.replace("libsql://", "https://") + "/v2/pipeline"
        self.headers = {
            "Authorization": f"Bearer {auth_token}",
            "Content-Type": "application/json",
        }

    def query(self, sql: str) -> list[dict]:
        """Execute a single SQL query, return rows as dicts."""
        resp = requests.post(
            self.url,
            headers=self.headers,
            json={"requests": [{"type": "execute", "stmt": {"sql": sql}}]},
            timeout=120,
        )
        resp.raise_for_status()
        data = resp.json()
        result = data["results"][0]["response"]["result"]
        cols = [c["name"] for c in result["cols"]]
        rows = []
        for row in result["rows"]:
            d = {}
            for i, col in enumerate(cols):
                val = row[i]
                d[col] = val.get("value") if isinstance(val, dict) else val
            rows.append(d)
        return rows


# ─── 数据获取 ──────────────────────────────────────────────────────────

def get_repo_ids(client: TursoClient) -> dict[str, int]:
    """Return {owner/repo: id} for all non-test repos."""
    rows = client.query(
        "SELECT id, owner, repo FROM repos WHERE owner NOT LIKE 'big-%' AND owner NOT LIKE 'perf-%' AND owner NOT LIKE 'test-%' ORDER BY id"
    )
    return {f"{r['owner']}/{r['repo']}": r["id"] for r in rows}


def fetch_runs(client: TursoClient, repo_ids: list[int], date_from: str, date_to: str) -> list[dict]:
    repo_id_list = ",".join(str(x) for x in repo_ids)
    # date 列格式不统一（有些带时间戳），用 LIKE 做前缀匹配
    return client.query(
        f"SELECT id, repo_id, name, head_branch, head_sha, event, "
        f"status, conclusion, created_at, updated_at, html_url, "
        f"duration_seconds, date "
        f"FROM runs "
        f"WHERE repo_id IN ({repo_id_list}) "
        f"AND (date LIKE '{date_from}%' OR date LIKE '{date_to}%' "
        f"     OR (date >= '{date_from}' AND date <= '{date_to}')) "
        f"ORDER BY created_at DESC"
    )


def fetch_jobs(client: TursoClient, run_ids: list[int]) -> list[dict]:
    all_jobs = []
    for i in range(0, len(run_ids), 5000):
        batch = run_ids[i : i + 5000]
        id_list = ",".join(str(x) for x in batch)
        jobs = client.query(
            f"SELECT id, run_id, name, status, conclusion, "
            f"created_at, started_at, completed_at, html_url, "
            f"queue_duration_seconds, duration_seconds "
            f"FROM jobs WHERE run_id IN ({id_list})"
        )
        all_jobs.extend(jobs)
    return all_jobs


def fetch_steps(client: TursoClient, job_ids: list[int]) -> list[dict]:
    """Fetch steps for given jobs with small batches to avoid Turso row limits."""
    all_steps = []
    # 每批 500 个 job_id，避免响应过大被截断
    batch_size = 500
    total = len(job_ids)
    for i in range(0, total, batch_size):
        batch = job_ids[i : i + batch_size]
        id_list = ",".join(str(x) for x in batch)
        steps = client.query(
            f"SELECT job_id, number, name, status, conclusion, "
            f"started_at, completed_at, duration_seconds "
            f"FROM steps WHERE job_id IN ({id_list})"
        )
        all_steps.extend(steps)
        if (i // batch_size + 1) % 20 == 0 or i + batch_size >= total:
            print(f"    Steps 进度: {min(i + batch_size, total)}/{total} jobs ({len(all_steps)} steps fetched)")
    return all_steps


def fetch_pr_metrics(client: TursoClient, repo_ids: list[int], date_from: str, date_to: str) -> list[dict]:
    repo_id_list = ",".join(str(x) for x in repo_ids)
    # created_at 格式: 2026-05-19T08:07:10.000+00:00
    return client.query(
        f"SELECT id, repo_id, pr_number, title, branch, author, state, "
        f"html_url, created_at, ci_started_at, ci_completed_at, "
        f"merged_at, time_to_ci_start_seconds, ci_duration_seconds, "
        f"time_to_merge_seconds, merge_lead_time_seconds, "
        f"workflow_count, successful_workflow_count, conclusion "
        f"FROM pr_metrics "
        f"WHERE repo_id IN ({repo_id_list}) "
        f"AND (created_at LIKE '{date_from}%' OR created_at LIKE '{date_to}%' "
        f"     OR (created_at >= '{date_from}' AND created_at <= '{date_to}T23:59:59')) "
        f"ORDER BY created_at DESC"
    )


def fetch_pr_workflows(client: TursoClient, pr_metric_ids: list[int]) -> list[dict]:
    if not pr_metric_ids:
        return []
    all_links = []
    for i in range(0, len(pr_metric_ids), 5000):
        batch = pr_metric_ids[i : i + 5000]
        id_list = ",".join(str(x) for x in batch)
        links = client.query(
            f"SELECT id, pr_metric_id, run_id "
            f"FROM pr_workflows WHERE pr_metric_id IN ({id_list})"
        )
        all_links.extend(links)
    return all_links


# ─── 统计工具 ──────────────────────────────────────────────────────────

def percentile(values: list[float], p: float) -> float:
    if not values:
        return 0.0
    s = sorted(values)
    k = (len(s) - 1) * (p / 100.0)
    f, c = math.floor(k), math.ceil(k)
    if f == c:
        return s[int(k)]
    return s[f] * (c - k) + s[c] * (k - f)


def sec_to_min(s) -> float | None:
    """Convert seconds to minutes, handling string values from DB."""
    if s is None:
        return None
    try:
        return round(float(s) / 60.0, 3)
    except (ValueError, TypeError):
        return None


def safe_div(a: float, b: float) -> float:
    return round(float(a) / float(b), 3) if b else 0.0


# ─── 分析逻辑 ──────────────────────────────────────────────────────────

def _sp(prefix: str | None, name: str) -> str:
    """Sheet name with optional repo prefix, sanitized for Excel."""
    if not prefix:
        return name[:31]
    safe_prefix = prefix.replace("/", "_").replace("\\\\", "").replace("*", "")
    max_prefix_len = 31 - len(name) - 3
    if max_prefix_len > 0:
        safe_prefix = safe_prefix[:max_prefix_len]
        return f"{safe_prefix} - {name}"
    return name[:31]


def _infer_resource_type(job_name: str) -> str:
    n = job_name.lower()
    if "310p" in n or "singlecard" in n or "single-card" in n:
        return "单卡 (310P)"
    if "multicard" in n or "multi-card" in n:
        if "4" in n:
            return "4卡"
        if "2" in n:
            return "2卡"
        return "多卡"
    if "nvidia" in n or "cuda" in n or "gpu" in n:
        return "NVIDIA GPU"
    if "e2e" in n:
        return "E2E"
    if "ut" in n or "unit" in n:
        return "单元测试"
    if "lint" in n or "check" in n:
        return "代码检查"
    return "其他"


def analyze_workflow_stats(runs, jobs):
    run_map = {r["id"]: r for r in runs}
    wf_groups = defaultdict(lambda: {"durations": [], "queues": [], "events": defaultdict(int)})

    for j in jobs:
        run = run_map.get(j["run_id"])
        if not run:
            continue
        wf = run["name"]
        dur = sec_to_min(j.get("duration_seconds"))
        if dur is not None:
            wf_groups[wf]["durations"].append(dur)
        q_dur = sec_to_min(j.get("queue_duration_seconds"))
        if q_dur is not None:
            wf_groups[wf]["queues"].append(q_dur)
        wf_groups[wf]["events"][run.get("event", "unknown")] += 1

    rows = []
    for wf, data in sorted(wf_groups.items(), key=lambda x: -len(x[1]["durations"])):
        durs = data["durations"]
        queues = data["queues"]
        dominant = max(data["events"], key=data["events"].get) if data["events"] else "unknown"
        rows.append({
            "工作流": wf,
            "触发类型": dominant,
            "运行次数": len(durs),
            "平均E2E(分钟)": safe_div(sum(durs), len(durs)),
            "P50 E2E(分钟)": percentile(durs, 50),
            "P90 E2E(分钟)": percentile(durs, 90),
            "平均排队(分钟)": safe_div(sum(queues), len(queues)),
            "P50 排队(分钟)": percentile(queues, 50),
            "P90 排队(分钟)": percentile(queues, 90),
            "调度周期(分钟)": round(max(durs) - min(durs), 3) if durs else 0,
        })
    return rows


def analyze_job_stats(runs, jobs):
    run_map = {r["id"]: r for r in runs}
    groups = defaultdict(lambda: {"durations": [], "queues": []})

    for j in jobs:
        run = run_map.get(j["run_id"])
        if not run:
            continue
        res_type = _infer_resource_type(j["name"])
        key = (f"{run['name']} / {j['name']}", res_type)
        d = sec_to_min(j.get("duration_seconds"))
        if d is not None:
            groups[key]["durations"].append(d)
        q = sec_to_min(j.get("queue_duration_seconds"))
        if q is not None:
            groups[key]["queues"].append(q)

    rows = []
    for (wf_job, res), data in sorted(groups.items(), key=lambda x: -len(x[1]["durations"])):
        durs = data["durations"]
        queues = data["queues"]
        rows.append({
            "工作流/任务": wf_job,
            "资源类型": res,
            "执行次数": len(durs),
            "平均E2E(分钟)": safe_div(sum(durs), len(durs)),
            "P50 E2E(分钟)": percentile(durs, 50),
            "P90 E2E(分钟)": percentile(durs, 90),
            "平均排队(分钟)": safe_div(sum(queues), len(queues)),
            "P50 排队(分钟)": percentile(queues, 50),
            "P90 排队(分钟)": percentile(queues, 90),
        })
    return rows


def analyze_step_stats(steps, step_names_map=None):
    groups = defaultdict(lambda: {"durations": [], "success": 0, "total": 0})

    for s in steps:
        name = s["name"]
        stype = step_names_map.get(name, "其他") if step_names_map else "其他"
        key = (name, stype)
        d = sec_to_min(s.get("duration_seconds"))
        if d is not None:
            groups[key]["durations"].append(d)
        groups[key]["total"] += 1
        if s.get("conclusion") == "success":
            groups[key]["success"] += 1

    rows = []
    for (name, stype), data in sorted(groups.items(), key=lambda x: -len(x[1]["durations"])):
        durs = data["durations"]
        rows.append({
            "步骤名称": name,
            "步骤类型": stype,
            "执行次数": data["total"],
            "平均耗时(分钟)": safe_div(sum(durs), len(durs)),
            "P50 耗时(分钟)": percentile(durs, 50),
            "P90 耗时(分钟)": percentile(durs, 90),
            "成功率": round(data["success"] / data["total"] * 100, 1) if data["total"] else 0,
        })
    return rows


def _calc_review_min(pm):
    if pm.get("merged_at") and pm.get("ci_completed_at"):
        try:
            merged = datetime.fromisoformat(pm["merged_at"].replace("Z", "+00:00"))
            ci_done = datetime.fromisoformat(pm["ci_completed_at"].replace("Z", "+00:00"))
            return sec_to_min((merged - ci_done).total_seconds())
        except (ValueError, TypeError):
            pass
    return None


def analyze_pr_stats(pr_metrics, pr_workflows):
    pr_wf_map = defaultdict(list)
    for pw in pr_workflows:
        pr_wf_map[pw["pr_metric_id"]].append(pw["run_id"])

    rows = []
    for pm in pr_metrics:
        wf_count = len(pr_wf_map.get(pm["id"], []))
        ci_dur = pm.get("ci_duration_seconds")
        rows.append({
            "PR编号": pm.get("pr_number"),
            "标题": pm.get("title", ""),
            "作者": pm.get("author", ""),
            "创建时间": pm.get("created_at", ""),
            "合并时间": pm.get("merged_at") or "",
            "CI完成时间": pm.get("ci_completed_at") or "",
            "PR E2E(分钟)": sec_to_min(ci_dur) if ci_dur else None,
            "CI后评审(分钟)": _calc_review_min(pm),
            "工作流数量": wf_count,
            "链接": pm.get("html_url", ""),
            "CI结论": pm.get("conclusion", ""),
        })
    return rows


def build_pr_details(pr_metrics, pr_workflows, runs, jobs, steps):
    run_map = {r["id"]: r for r in runs}
    run_jobs = defaultdict(list)
    for j in jobs:
        run_jobs[j["run_id"]].append(j)
    job_steps = defaultdict(list)
    for s in steps:
        job_steps[s["job_id"]].append(s)
    pr_wf_map = defaultdict(list)
    for pw in pr_workflows:
        pr_wf_map[pw["pr_metric_id"]].append(pw["run_id"])

    _COMMON_KEYS = [
        "层级", "PR编号", "PR标题", "PR作者", "PR创建时间", "PR合并时间",
        "PR E2E(分钟)", "CI后评审(分钟)",
        "工作流名称", "工作流运行ID", "工作流状态", "工作流结论",
        "工作流创建时间", "工作流开始时间", "工作流完成时间", "工作流耗时(分钟)",
        "任务名称", "任务ID", "任务状态", "任务结论",
        "任务创建时间", "任务开始时间", "任务完成时间",
        "任务排队(分钟)", "任务耗时(分钟)",
        "步骤序号", "步骤名称", "步骤状态", "步骤结论",
        "步骤开始时间", "步骤完成时间", "步骤耗时(分钟)",
        "链接",
    ]

    def _base(level, pm, pr_e2e, review):
        return {k: None for k in _COMMON_KEYS} | {
            "层级": level,
            "PR编号": pm.get("pr_number"),
            "PR标题": pm.get("title", ""),
            "PR作者": pm.get("author", ""),
            "PR创建时间": pm.get("created_at", ""),
            "PR合并时间": pm.get("merged_at") or "",
            "PR E2E(分钟)": pr_e2e,
            "CI后评审(分钟)": review,
        }

    all_rows = []
    for pm in pr_metrics:
        pr_e2e = sec_to_min(pm.get("ci_duration_seconds")) if pm.get("ci_duration_seconds") else None
        review = _calc_review_min(pm)
        wf_run_ids = pr_wf_map.get(pm["id"], [])

        all_rows.append(_base("PR", pm, pr_e2e, review) | {"链接": pm.get("html_url", "")})

        for run_id in wf_run_ids:
            run = run_map.get(run_id)
            if not run:
                continue
            wf_dur = sec_to_min(run.get("duration_seconds"))
            all_rows.append(_base("WORKFLOW", pm, pr_e2e, review) | {
                "工作流名称": run.get("name"),
                "工作流运行ID": run_id,
                "工作流状态": run.get("status"),
                "工作流结论": run.get("conclusion"),
                "工作流创建时间": run.get("created_at"),
                "工作流开始时间": run.get("created_at"),
                "工作流完成时间": run.get("updated_at"),
                "工作流耗时(分钟)": wf_dur,
                "链接": run.get("html_url", ""),
            })

            for j in run_jobs.get(run_id, []):
                all_rows.append(_base("JOB", pm, pr_e2e, review) | {
                    "工作流名称": run.get("name"),
                    "工作流运行ID": run_id,
                    "工作流状态": run.get("status"),
                    "工作流结论": run.get("conclusion"),
                    "工作流创建时间": run.get("created_at"),
                    "工作流开始时间": run.get("created_at"),
                    "工作流完成时间": run.get("updated_at"),
                    "工作流耗时(分钟)": wf_dur,
                    "任务名称": j.get("name"),
                    "任务ID": j.get("id"),
                    "任务状态": j.get("status"),
                    "任务结论": j.get("conclusion"),
                    "任务创建时间": j.get("created_at"),
                    "任务开始时间": j.get("started_at"),
                    "任务完成时间": j.get("completed_at"),
                    "任务排队(分钟)": sec_to_min(j.get("queue_duration_seconds")),
                    "任务耗时(分钟)": sec_to_min(j.get("duration_seconds")),
                    "链接": j.get("html_url", ""),
                })

                for s in sorted(job_steps.get(j["id"], []), key=lambda x: x.get("number", 0)):
                    all_rows.append(_base("STEP", pm, pr_e2e, review) | {
                        "工作流名称": run.get("name"),
                        "工作流运行ID": run_id,
                        "工作流状态": run.get("status"),
                        "工作流结论": run.get("conclusion"),
                        "工作流创建时间": run.get("created_at"),
                        "工作流开始时间": run.get("created_at"),
                        "工作流完成时间": run.get("updated_at"),
                        "工作流耗时(分钟)": wf_dur,
                        "任务名称": j.get("name"),
                        "任务ID": j.get("id"),
                        "任务状态": j.get("status"),
                        "任务结论": j.get("conclusion"),
                        "任务创建时间": j.get("created_at"),
                        "任务开始时间": j.get("started_at"),
                        "任务完成时间": j.get("completed_at"),
                        "任务排队(分钟)": sec_to_min(j.get("queue_duration_seconds")),
                        "任务耗时(分钟)": sec_to_min(j.get("duration_seconds")),
                        "步骤序号": s.get("number"),
                        "步骤名称": s.get("name"),
                        "步骤状态": s.get("status"),
                        "步骤结论": s.get("conclusion"),
                        "步骤开始时间": s.get("started_at"),
                        "步骤完成时间": s.get("completed_at"),
                        "步骤耗时(分钟)": sec_to_min(s.get("duration_seconds")),
                    })

    return all_rows


def analyze_comparison(repos_data: dict[str, dict]) -> list[dict]:
    rows = []
    for repo_name, data in repos_data.items():
        runs = data.get("runs", [])
        jobs = data.get("jobs", [])
        pr_metrics = data.get("pr_metrics", [])

        wf_durs = [float(r["duration_seconds"]) for r in runs if r.get("duration_seconds") is not None]
        job_durs = [float(j["duration_seconds"]) for j in jobs if j.get("duration_seconds") is not None]
        job_queues = [float(j.get("queue_duration_seconds", 0) or 0) for j in jobs]

        conclusions = defaultdict(int)
        for j in jobs:
            conclusions[j.get("conclusion") or "unknown"] += 1
        total = len(jobs)

        merged_prs = sum(1 for pm in pr_metrics if pm.get("merged_at"))
        ci_durations = [float(pm["ci_duration_seconds"]) for pm in pr_metrics if pm.get("ci_duration_seconds") is not None]

        events = defaultdict(int)
        for r in runs:
            events[r.get("event") or "unknown"] += 1

        rows.append({
            "仓库": repo_name,
            "总 Run 数": len(runs),
            "总 Job 数": total,
            "平均 Run 耗时(分钟)": safe_div(sum(wf_durs), len(wf_durs)) if wf_durs else 0,
            "P50 Run 耗时(分钟)": percentile(wf_durs, 50) if wf_durs else 0,
            "P90 Run 耗时(分钟)": percentile(wf_durs, 90) if wf_durs else 0,
            "平均 Job 耗时(分钟)": safe_div(sum(job_durs), len(job_durs)) if job_durs else 0,
            "P50 Job 耗时(分钟)": percentile(job_durs, 50) if job_durs else 0,
            "P90 Job 耗时(分钟)": percentile(job_durs, 90) if job_durs else 0,
            "平均排队(分钟)": safe_div(sum(job_queues), len(job_queues)) if job_queues else 0,
            "Job 成功率": round(conclusions.get("success", 0) / total * 100, 1) if total else 0,
            "Job 失败率": round(conclusions.get("failure", 0) / total * 100, 1) if total else 0,
            "PR 数量": len(pr_metrics),
            "已合并 PR": merged_prs,
            "平均 CI 时长(分钟)": safe_div(sum(ci_durations), len(ci_durations)) if ci_durations else 0,
            "主要触发类型": max(events, key=events.get) if events else "N/A",
        })
    return rows


# ─── Excel 输出 ─────────────────────────────────────────────────────────

def write_excel(filepath: str, sheets: dict[str, list[dict]]):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠ openpyxl 未安装，跳过 Excel 输出")
        print("    安装: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])
        headers = list(rows[0].keys())

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        for ri, row_data in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=row_data.get(h))
                cell.border = thin_border
                if ri % 2 == 0:
                    cell.fill = alt_fill

        for ci in range(1, len(headers) + 1):
            max_len = len(str(headers[ci - 1]))
            for row in range(2, min(len(rows) + 2, 100)):
                val = ws.cell(row=row, column=ci).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 55)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    print(f"  ✅ 已保存: {filepath}")


# ─── 终端输出 ──────────────────────────────────────────────────────────

def print_summary(repos_data: dict[str, dict]):
    for repo_name, data in repos_data.items():
        runs = data.get("runs", [])
        jobs = data.get("jobs", [])
        pr_metrics = data.get("pr_metrics", [])

        conclusions = defaultdict(int)
        for j in jobs:
            conclusions[j.get("conclusion") or "unknown"] += 1
        total = len(jobs)

        job_durs = [float(j["duration_seconds"]) for j in jobs if j.get("duration_seconds") is not None]
        job_queues = [float(j.get("queue_duration_seconds", 0) or 0) for j in jobs]

        print(f"\n{'='*60}")
        print(f"📊 {repo_name}")
        print(f"{'='*60}")
        print(f"  Runs:          {len(runs)}")
        print(f"  Jobs:          {total}")
        print(f"  PRs:           {len(pr_metrics)}")
        if total:
            print(f"  Job 成功率:    {conclusions.get('success', 0)}/{total} ({round(conclusions['success']/total*100, 1)}%)")
            print(f"  Job 失败率:    {conclusions.get('failure', 0)}/{total} ({round(conclusions.get('failure',0)/total*100, 1)}%)")
        if job_durs:
            avg_dur = sum(job_durs)/len(job_durs)
            print(f"  Job 平均耗时:  {avg_dur/60:.1f} 分钟")
            print(f"  Job P50 耗时:  {percentile(job_durs, 50)/60:.1f} 分钟")
            print(f"  Job P90 耗时:  {percentile(job_durs, 90)/60:.1f} 分钟")
        if job_queues:
            avg_q = sum(job_queues)/len(job_queues)
            print(f"  平均排队:      {avg_q/60:.1f} 分钟")
            print(f"  P90 排队:      {percentile(job_queues, 90)/60:.1f} 分钟")

        # Top 5 slowest jobs
        if job_durs:
            sorted_jobs = sorted(jobs, key=lambda j: float(j.get("duration_seconds", 0) or 0), reverse=True)[:5]
            print(f"\n  最慢的 5 个 Job:")
            for j in sorted_jobs:
                d = float(j.get("duration_seconds", 0) or 0) / 60
                print(f"    {j['name']}: {d:.1f} 分钟")

        # Top 5 workflows by run count
        wf_counts = defaultdict(int)
        for r in runs:
            wf_counts[r["name"]] += 1
        print(f"\n  最频繁的 5 个 Workflow:")
        for wf, cnt in sorted(wf_counts.items(), key=lambda x: -x[1])[:5]:
            print(f"    {wf}: {cnt} 次")


# ─── 主流程 ─────────────────────────────────────────────────────────────

def fetch_all_for_repo(client: TursoClient, repo_id: int, date_from: str, date_to: str):
    """Fetch all data for a single repo."""
    runs = fetch_runs(client, [repo_id], date_from, date_to)
    if not runs:
        return {"runs": [], "jobs": [], "steps": [], "pr_metrics": [], "pr_workflows": []}

    run_ids = [r["id"] for r in runs]
    jobs = fetch_jobs(client, run_ids)
    job_ids = [j["id"] for j in jobs]
    steps = fetch_steps(client, job_ids)

    pr_metrics = fetch_pr_metrics(client, [repo_id], date_from, date_to)
    pr_ids = [pm["id"] for pm in pr_metrics]
    pr_workflows = fetch_pr_workflows(client, pr_ids)

    return {
        "runs": runs,
        "jobs": jobs,
        "steps": steps,
        "pr_metrics": pr_metrics,
        "pr_workflows": pr_workflows,
    }


def main():
    parser = argparse.ArgumentParser(description="CI 效率分析报告生成器")
    parser.add_argument("--repo", action="append", help="仓库名 owner/repo（可多次指定）")
    parser.add_argument("--from", dest="date_from", help="起始日期 YYYY-MM-DD")
    parser.add_argument("--to", dest="date_to", help="结束日期 YYYY-MM-DD")
    parser.add_argument("--list-repos", action="store_true", help="列出所有可用仓库")
    parser.add_argument("--step-names", help="Step 分类映射 JSON 文件路径")
    parser.add_argument("--no-excel", action="store_true", help="跳过 Excel 输出")
    parser.add_argument("--skip-steps", action="store_true", help="跳过 steps 数据（加速查询）")
    parser.add_argument("--output", "-o", help="输出 Excel 文件路径（默认自动生成）")
    args = parser.parse_args()

    # Load env
    env = load_env(ENV_FILE)
    db_url = env.get("TURSO_DATABASE_URL")
    auth_token = env.get("TURSO_AUTH_TOKEN")
    if not db_url or not auth_token:
        print("❌ 未找到 TURSO_DATABASE_URL 或 TURSO_AUTH_TOKEN，请检查 .env 文件")
        sys.exit(1)

    client = TursoClient(db_url, auth_token)

    # Load step names map
    step_names_path = Path(args.step_names) if args.step_names else DEFAULT_STEP_NAMES
    step_names_map = {}
    if step_names_path.exists():
        with open(step_names_path) as f:
            step_names_map = json.load(f)
        print(f"📂 加载 step 分类: {step_names_path} ({len(step_names_map)} 条映射)")

    # List repos mode
    if args.list_repos:
        repos = get_repo_ids(client)
        print(f"\n📋 可用仓库 (共 {len(repos)} 个):")
        for name, rid in sorted(repos.items(), key=lambda x: x[1]):
            print(f"  {rid:>5}: {name}")
        return

    # Resolve repos
    all_repos = get_repo_ids(client)
    if args.repo:
        repo_map = {}
        for r in args.repo:
            if r in all_repos:
                repo_map[r] = all_repos[r]
            else:
                print(f"⚠ 未找到仓库: {r}")
                print(f"  可用仓库: {', '.join(list(all_repos.keys())[:10])}...")
                sys.exit(1)
    else:
        # 默认使用 vllm-ascend
        default_repo = "vllm-project/vllm-ascend"
        if default_repo in all_repos:
            repo_map = {default_repo: all_repos[default_repo]}
        else:
            # 使用第一个可用仓库
            first = list(all_repos.items())[0]
            repo_map = {first[0]: first[1]}
            print(f"  默认仓库: {first[0]}")

    # Date range
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    date_from = args.date_from or (datetime.now(timezone.utc) - timedelta(days=DEFAULT_DAYS)).strftime("%Y-%m-%d")
    date_to = args.date_to or today
    print(f"\n📅 时间范围: {date_from} → {date_to}")
    print(f"📦 仓库: {', '.join(repo_map.keys())}")

    # Fetch data
    repos_data = {}
    for repo_name, repo_id in repo_map.items():
        print(f"\n⏳ 获取 {repo_name} (id={repo_id}) 数据...")
        data = fetch_all_for_repo(client, repo_id, date_from, date_to)
        repos_data[repo_name] = data
        print(f"  ✅ Runs: {len(data['runs'])}, Jobs: {len(data['jobs'])}, "
              f"Steps: {len(data['steps'])}, PRs: {len(data['pr_metrics'])}")

    if not any(d["runs"] for d in repos_data.values()):
        print("\n⚠ 指定时间范围内没有数据")
        return

    # Terminal summary
    print_summary(repos_data)

    # Build analysis sheets
    sheets = {}

    # Multi-repo comparison
    if len(repo_map) > 1:
        sheets["仓库对比"] = analyze_comparison(repos_data)

    # Per-repo analysis
    for repo_name, data in repos_data.items():
        runs, jobs, steps = data["runs"], data["jobs"], data["steps"]
        pr_metrics, pr_workflows = data["pr_metrics"], data["pr_workflows"]

        sheet_prefix = repo_name if len(repo_map) > 1 else None

        if len(runs) == 0:
            continue

        wf = analyze_workflow_stats(runs, jobs)
        sheets[_sp(sheet_prefix, "工作流统计")] = wf

        js = analyze_job_stats(runs, jobs)
        sheets[_sp(sheet_prefix, "任务统计")] = js

        if steps:
            ss = analyze_step_stats(steps, step_names_map)
            sheets[_sp(sheet_prefix, "步骤统计")] = ss

        ps = analyze_pr_stats(pr_metrics, pr_workflows)
        sheets[_sp(sheet_prefix, "PR统计")] = ps

        pd_rows = build_pr_details(pr_metrics, pr_workflows, runs, jobs, steps)
        sheets[_sp(sheet_prefix, "PR详情")] = pd_rows

    # Write Excel
    if not args.no_excel and sheets:
        if args.output:
            outfile = args.output
        else:
            date_tag = f"{date_from}_to_{date_to}"
            repo_tag = "_vs_".join(r.replace("/", "_") for r in repo_map.keys())
            outfile = f"{repo_tag}-ci-report-{date_tag}.xlsx"
        write_excel(outfile, sheets)
        print(f"\n📊 共生成 {len(sheets)} 个 sheet: {', '.join(sheets.keys())}")


if __name__ == "__main__":
    main()