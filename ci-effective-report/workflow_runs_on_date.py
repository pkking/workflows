#!/usr/bin/env python3
"""按"运行创建日期"分析某个 GitHub Actions workflow 的耗时（run / job / step）。

与 github_ci_efficiency_report.py（以 PR 合并日期为口径、从 PR head SHA 反查 run）不同，
本脚本直接按 run 的 created_at 日期列某 workflow 的所有 run，因此能覆盖定时/push/PR 等
所有触发类型在指定日期的运行。

复用现有脚本的 GitHubClient / 数据类 / job+step 解析。

用法:
  export GITHUB_TOKEN=$(gh auth token)
  python3 workflow_runs_on_date.py \
    --repo vllm-project/vllm-ascend --workflow E2E --date 2026-06-29 \
    --output e2e-2026-06-29.xlsx
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

# 复用现有 Method B 脚本的客户端与数据类
SCRIPT_DIR = Path(__file__).parent
SKILL_SCRIPTS = SCRIPT_DIR / "skills" / "github-ci-efficiency-report" / "scripts"
sys.path.insert(0, str(SKILL_SCRIPTS))
from github_ci_efficiency_report import (  # noqa: E402
    GitHubClient,
    WorkflowRunInfo,
    JobInfo,
    parse_dt,
    minutes_between,
    percentile,
    average,
    write_sheet,
    workflow_matches,
)


def fetch_workflow_id(client: GitHubClient, repo: str, workflow_pattern: str) -> tuple[int, str] | None:
    """按 workflow 显示名匹配，返回 (id, name)。优先精确名匹配，否则子串。"""
    workflows: list[dict] = []
    page = 1
    while True:
        data, _ = client.get(f"/repos/{repo}/actions/workflows", {"per_page": 100, "page": page})
        ws = data.get("workflows", [])
        if not ws:
            break
        workflows.extend(ws)
        if len(workflows) >= data.get("total_count", len(workflows)):
            break
        page += 1
    exact = [w for w in workflows if w.get("name") == workflow_pattern and w.get("state") == "active"]
    matches = exact or [w for w in workflows if workflow_matches(w.get("name", ""), [workflow_pattern]) and w.get("state") == "active"]
    if not matches:
        return None
    if len(matches) > 1:
        names = ", ".join(f'{w["name"]}({w["id"]})' for w in matches)
        print(f"WARNING: 多个 workflow 命中 '{workflow_pattern}': {names}; 取第一个 {matches[0]['name']}", file=sys.stderr)
    w = matches[0]
    return int(w["id"]), w["name"]


def fetch_runs_on_date(client: GitHubClient, repo: str, workflow_id: int, date: str) -> list[WorkflowRunInfo]:
    """列出某 workflow 在指定日期(UTC)创建的所有 run。created=YYYY-MM-DD = 当天全天。"""
    runs: list[WorkflowRunInfo] = []
    params = {"created": date, "per_page": 100}
    next_params = dict(params)
    page = 1
    while True:
        p = dict(next_params)
        if page > 1:
            p["page"] = page
        data, headers = client.get(f"/repos/{repo}/actions/workflows/{workflow_id}/runs", p)
        items = data.get("workflow_runs", [])
        if not items:
            break
        for run in items:
            runs.append(WorkflowRunInfo(
                id=int(run["id"]),
                name=run.get("name") or str(run.get("workflow_id") or ""),
                workflow_id=run.get("workflow_id"),
                status=run.get("status", ""),
                conclusion=run.get("conclusion") or "",
                event=run.get("event", ""),
                html_url=run.get("html_url", ""),
                head_sha=run.get("head_sha", ""),
                created_at=parse_dt(run.get("created_at")),
                run_started_at=parse_dt(run.get("run_started_at")),
                updated_at=parse_dt(run.get("updated_at")),
            ))
        page += 1
        # GitHub runs 列表用 Link 头分页；无 next 则止
        if not has_next(headers.get("link", "")):
            break
    return runs


def has_next(link_header: str) -> bool:
    return 'rel="next"' in link_header


def fetch_jobs(client: GitHubClient, repo: str, run_id: int) -> list[JobInfo]:
    """复用脚本同名逻辑（jobs 含 steps），内联以避免 import 依赖签名变化。"""
    from github_ci_efficiency_report import fetch_jobs_for_run
    return fetch_jobs_for_run(client, repo, run_id)


# ─── job 名归一化（矩阵变体/分片去除） ───────────────────────────────

def norm_job_name(name: str) -> str:
    """run-selected-tests (v0.23.0) / a3-4 card-(part 2-3) -> run-selected-tests / a3-4"""
    s = re.sub(r'\s*card-\([^)]*\)', '', name)      # 去 card-(part X-Y)
    s = re.sub(r'\s*\([^)]*\)', '', s)               # 去矩阵变体 (v0.23.0) / (commit)
    s = re.sub(r'\s+', ' ', s).strip(' /')
    parts = [p.strip() for p in s.split('/') if p.strip()]
    if len(parts) >= 2:
        return f"{parts[0]} / {parts[-1]}"
    return parts[0] if parts else name


# ─── 统计 ────────────────────────────────────────────────────────────

def build_run_stats(runs: list[WorkflowRunInfo]) -> list[list]:
    by_conc = defaultdict(list)
    for r in runs:
        if r.duration_min is not None:
            by_conc[r.conclusion or "unknown"].append(r.duration_min)
    rows = [["结论", "运行数", "平均E2E(分钟)", "P50(分钟)", "P90(分钟)", "总耗时(分钟)"]]
    for conc in sorted(by_conc):
        v = by_conc[conc]
        rows.append([conc, len(v), average(v), percentile(v, 0.50), percentile(v, 0.90), round(sum(v), 1)])
    allv = [r.duration_min for r in runs if r.duration_min is not None]
    rows.append([])
    rows.append(["合计", len(runs), average(allv), percentile(allv, 0.50), percentile(allv, 0.90), round(sum(allv), 1)])
    return rows


def build_run_details(runs, run_jobs) -> list[list]:
    rows = [["运行ID", "创建时间", "事件", "结论", "E2E(分钟)", "Job数", "链接"]]
    for r in sorted(runs, key=lambda x: x.created_at or parse_dt("2000-01-01")):
        rows.append([r.id, (r.created_at.isoformat(timespec="seconds") if r.created_at else ""),
                     r.event, r.conclusion, r.duration_min, len(run_jobs.get(r.id, [])), r.html_url])
    return rows


def build_job_stats(run_jobs) -> list[list]:
    groups: dict[str, dict] = defaultdict(lambda: {"dur": [], "queue": [], "succ": 0, "total": 0})
    for jobs in run_jobs.values():
        for j in jobs:
            if j.conclusion not in ("success", "failure"):
                continue  # 跳过 skipped/cancelled（无真实执行时间）
            key = norm_job_name(j.name)
            if j.duration_min is not None:
                groups[key]["dur"].append(j.duration_min)
            if j.queue_min is not None:
                groups[key]["queue"].append(j.queue_min)
            groups[key]["total"] += 1
            if j.conclusion == "success":
                groups[key]["succ"] += 1
    rows = [["任务(归一化)", "执行次数", "成功", "成功率",
             "平均E2E(分钟)", "P50(分钟)", "P90(分钟)",
             "平均排队(分钟)", "P50排队(分钟)", "P90排队(分钟)", "总耗时(分钟)"]]
    for key in sorted(groups, key=lambda k: -sum(groups[k]["dur"])):
        g = groups[key]
        d, q = g["dur"], g["queue"]
        rows.append([key, g["total"], g["succ"],
                     round(g["succ"] / g["total"], 3) if g["total"] else None,
                     average(d), percentile(d, 0.50), percentile(d, 0.90),
                     average(q), percentile(q, 0.50), percentile(q, 0.90), round(sum(d), 1)])
    return rows


def build_step_stats(run_jobs) -> list[list]:
    dur: dict[str, list[float]] = defaultdict(list)
    conc: dict[str, list[str]] = defaultdict(list)
    for jobs in run_jobs.values():
        for j in jobs:
            if j.conclusion not in ("success", "failure"):
                continue
            for s in j.steps:
                if s.duration_min is None or s.duration_min <= 0:
                    continue
                dur[s.name].append(s.duration_min)
                conc[s.name].append(s.conclusion)
    rows = [["步骤名称", "执行次数", "成功率",
             "平均耗时(分钟)", "P50(分钟)", "P90(分钟)", "总耗时(分钟)"]]
    for name in sorted(dur, key=lambda k: -sum(dur[k])):
        d = dur[name]
        c = conc[name]
        succ = sum(1 for x in c if x == "success")
        rows.append([name, len(d), round(succ / len(c), 3) if c else None,
                     average(d), percentile(d, 0.50), percentile(d, 0.90), round(sum(d), 1)])
    return rows


def build_job_details(runs, run_jobs) -> list[list]:
    rmap = {r.id: r for r in runs}
    rows = [["运行ID", "运行结论", "任务(原始)", "任务(归一化)", "任务结论",
             "排队(分钟)", "E2E(分钟)", "步骤数", "链接"]]
    items = []
    for rid, jobs in run_jobs.items():
        r = rmap.get(rid)
        for j in jobs:
            items.append((r, j))
    items.sort(key=lambda x: -(x[1].duration_min or 0))
    for r, j in items[:500]:  # ponytail: 最慢 500 个 job，全量见 job_stats
        rows.append([r.id if r else None,
                     r.conclusion if r else "", j.name, norm_job_name(j.name), j.conclusion,
                     j.queue_min, j.duration_min, len(j.steps),
                     r.html_url if r else ""])
    return rows


# ─── 主流程 ──────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="按运行创建日期分析某 workflow 的 run/job/step 耗时")
    ap.add_argument("--repo", required=True, help="owner/name")
    ap.add_argument("--workflow", required=True, help="workflow 显示名子串（如 E2E）")
    ap.add_argument("--date", required=True, help="YYYY-MM-DD（UTC 当天全天）")
    ap.add_argument("--output", help="输出 xlsx 路径")
    ap.add_argument("--token", default=os.getenv("GITHUB_TOKEN") or os.getenv("GH_TOKEN"))
    ap.add_argument("--concurrency", type=int, default=8)
    args = ap.parse_args()
    if not args.token:
        print("需要 --token 或 GITHUB_TOKEN/GH_TOKEN（可用 gh auth token）", file=sys.stderr)
        return 2
    if not args.output:
        safe = args.repo.replace("/", "_")
        args.output = f"{safe}-{args.workflow}-{args.date}.xlsx"

    client = GitHubClient(args.token)
    wf = fetch_workflow_id(client, args.repo, args.workflow)
    if not wf:
        print(f"未找到名称含 '{args.workflow}' 的 active workflow", file=sys.stderr)
        return 1
    wid, wname = wf
    print(f"workflow: {wname} (id={wid})", file=sys.stderr)

    t0 = time.time()
    runs = fetch_runs_on_date(client, args.repo, wid, args.date)
    print(f"{args.date} (UTC) 共 {len(runs)} 个 run（API 调用 {client.calls}，{time.time()-t0:.1f}s）", file=sys.stderr)
    if not runs:
        return 1

    # 并发抓 jobs（含 steps）
    print(f"并发抓取 {len(runs)} 个 run 的 jobs（concurrency={args.concurrency}）...", file=sys.stderr)
    run_jobs: dict[int, list[JobInfo]] = {}
    failed = []

    def one(rid):
        try:
            return rid, fetch_jobs(client, args.repo, rid)
        except Exception as e:
            print(f"WARNING: run {rid} jobs 失败: {e}", file=sys.stderr)
            return rid, []

    with ThreadPoolExecutor(max_workers=args.concurrency) as ex:
        for fut in as_completed({ex.submit(one, r.id): r.id for r in runs}):
            rid, jobs = fut.result()
            run_jobs[rid] = jobs
            if not jobs:
                failed.append(rid)
    print(f"jobs 抓取完成: API 调用 {client.calls}，{time.time()-t0:.1f}s，{len(failed)} 个 run 无 job", file=sys.stderr)

    # 终端摘要
    njobs = sum(len(v) for v in run_jobs.values())
    nsteps = sum(len(s.steps) for v in run_jobs.values() for s in v)
    concl = defaultdict(int)
    for r in runs:
        concl[r.conclusion or "unknown"] += 1
    print(f"\n{'='*60}")
    print(f"📊 {args.repo} / {wname}  —  {args.date} (UTC)")
    print(f"{'='*60}")
    print(f"  Runs:  {len(runs)}  ({dict(concl)})")
    print(f"  Jobs:  {njobs}   Steps: {nsteps}")
    rdur = [r.duration_min for r in runs if r.duration_min is not None]
    if rdur:
        print(f"  Run E2E: 平均 {average(rdur)} / P50 {percentile(rdur,0.5)} / P90 {percentile(rdur,0.9)} 分钟, 总 {round(sum(rdur),1)} 分钟")

    # 最慢 10 job（归一化）
    jg: dict[str, list[float]] = defaultdict(list)
    for jobs in run_jobs.values():
        for j in jobs:
            if j.conclusion in ("success", "failure") and j.duration_min:
                jg[norm_job_name(j.name)].append(j.duration_min)
    print(f"\n  最慢 10 个任务(归一化, 按总耗时):")
    for k, v in sorted(jg.items(), key=lambda x: -sum(x[1]))[:10]:
        print(f"    {k}: {len(v)}次, 平均 {average(v)}/P90 {percentile(v,0.9)} 分钟, 总 {round(sum(v),1)} 分钟")

    # 最慢 10 step
    sg: dict[str, list[float]] = defaultdict(list)
    for jobs in run_jobs.values():
        for j in jobs:
            if j.conclusion not in ("success", "failure"):
                continue
            for s in j.steps:
                if s.duration_min and s.duration_min > 0:
                    sg[s.name].append(s.duration_min)
    print(f"\n  最慢 10 个步骤(按总耗时):")
    for k, v in sorted(sg.items(), key=lambda x: -sum(x[1]))[:10]:
        print(f"    {k}: {len(v)}次, 平均 {average(v)}/P90 {percentile(v,0.9)} 分钟, 总 {round(sum(v),1)} 分钟")

    # Excel
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl 未安装，仅终端输出", file=sys.stderr)
        return 0
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("run_stats"); write_sheet(ws, build_run_stats(runs))
    ws = wb.create_sheet("run_details"); write_sheet(ws, build_run_details(runs, run_jobs))
    ws = wb.create_sheet("job_stats"); write_sheet(ws, build_job_stats(run_jobs))
    ws = wb.create_sheet("step_stats"); write_sheet(ws, build_step_stats(run_jobs))
    ws = wb.create_sheet("job_details(top500)"); write_sheet(ws, build_job_details(runs, run_jobs))
    wb.save(args.output)
    print(f"\n✅ 已保存: {args.output}", file=sys.stderr)
    print(f"   API 调用: {client.calls}，耗时 {time.time()-t0:.1f}s", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
