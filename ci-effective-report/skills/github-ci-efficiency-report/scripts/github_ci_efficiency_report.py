#!/usr/bin/env python3
"""Export GitHub PR and Actions CI efficiency metrics to an Excel workbook."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import socket
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from typing import Any


API_ROOT = "https://api.github.com"


def load_step_types(path: str) -> dict[str, str]:
    """Load step type classification from a JSON file. Returns {step_name: category}."""
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def export_step_names(prs: list[PullRequestInfo], output: str) -> None:
    """Collect unique step names and write to JSON for external LLM classification."""
    names: set[str] = set()
    for pr in prs:
        for run in pr.workflows:
            for job in run.jobs:
                for step in job.steps:
                    if step.name:
                        names.add(step.name)
    with open(output, "w", encoding="utf-8") as f:
        json.dump({name: "" for name in sorted(names)}, f, ensure_ascii=False, indent=2)
    print(f"Exported {len(names)} unique step names to {output}. Fill in the values and pass back via --step-types.", file=sys.stderr)


def parse_dt(value: str | None) -> dt.datetime | None:
    if not value:
        return None
    if value.endswith("Z"):
        value = value[:-1] + "+00:00"
    return dt.datetime.fromisoformat(value).astimezone(dt.timezone.utc)


def fmt_dt(value: dt.datetime | None) -> str:
    return value.isoformat(timespec="seconds").replace("+00:00", "Z") if value else ""


def minutes_between(start: dt.datetime | None, end: dt.datetime | None, allow_negative: bool = False) -> float | None:
    if not start or not end:
        return None
    value = round((end - start).total_seconds() / 60.0, 3)
    if value < 0 and not allow_negative:
        return None
    return value


def percentile(values: list[float], pct: float) -> float | None:
    clean = sorted(v for v in values if v is not None)
    if not clean:
        return None
    if len(clean) == 1:
        return round(clean[0], 3)
    rank = (len(clean) - 1) * pct
    lower = math.floor(rank)
    upper = math.ceil(rank)
    if lower == upper:
        return round(clean[lower], 3)
    weight = rank - lower
    return round(clean[lower] * (1 - weight) + clean[upper] * weight, 3)


def average(values: list[float]) -> float | None:
    clean = [v for v in values if v is not None]
    return round(sum(clean) / len(clean), 3) if clean else None


# --- Trigger type mapping ---

EVENT_TO_TRIGGER_TYPE = {
    "pull_request": "PR触发",
    "pull_request_target": "PR触发",
    "push": "push触发",
    "schedule": "定时触发",
    "workflow_dispatch": "人工触发",
    "workflow_call": "人工触发",
    "release": "release触发",
    "issue_comment": "issue_comment触发",
    "create": "create触发",
}


def get_trigger_type(event: str) -> str:
    return EVENT_TO_TRIGGER_TYPE.get(event, event or "unknown")


COMMON_LABELS = {"self-hosted", "linux", "x64", "arm64", "windows", "macos", "ubuntu", "ubuntu-latest", "ubuntu-22.04", "ubuntu-20.04", "ubuntu-24.04"}


def _extract_resource_type(labels: list[str]) -> str:
    for label in labels:
        if label.lower() not in COMMON_LABELS:
            return label
    return ",".join(labels) if labels else ""


# --- Cron schedule interval parser ---

def cron_to_interval_minutes(cron_expr: str) -> float | None:
    """Parse a standard cron expression and return the approximate interval in minutes."""
    parts = cron_expr.strip().split()
    if len(parts) != 5:
        return None
    minute, hour, dom, month, dow = parts

    # */N * * * *  → every N minutes
    if minute.startswith("*/") and hour == "*" and dom == "*" and month == "*" and dow == "*":
        try:
            return float(minute[2:])
        except ValueError:
            return None

    # 0 */N * * *  → every N hours
    if minute == "0" and hour.startswith("*/") and dom == "*" and month == "*" and dow == "*":
        try:
            return float(hour[2:]) * 60
        except ValueError:
            return None

    # 0 0 * * *  → daily
    if minute == "0" and hour == "0" and dom == "*" and month == "*" and dow == "*":
        return 1440.0

    # 0 * * * *  → hourly
    if minute == "0" and hour == "*" and dom == "*" and month == "*" and dow == "*":
        return 60.0

    # 0 0 * * N  → weekly (specific day of week)
    if minute == "0" and hour == "0" and dom == "*" and month == "*" and dow != "*":
        return 10080.0

    # 0 0 N * *  → monthly (specific day of month)
    if minute == "0" and hour == "0" and dom != "*" and month == "*" and dow == "*":
        return 43200.0

    # Fallback: return None for complex expressions
    return None


def parse_schedule_interval(cron_expressions: list[str]) -> float | None:
    """Given a list of cron expressions from a workflow, return the minimum interval in minutes."""
    intervals = []
    for expr in cron_expressions:
        interval = cron_to_interval_minutes(expr)
        if interval is not None:
            intervals.append(interval)
    return min(intervals) if intervals else None


@dataclass
class StepInfo:
    name: str
    status: str
    conclusion: str
    number: int | None
    started_at: dt.datetime | None
    completed_at: dt.datetime | None

    @property
    def duration_min(self) -> float | None:
        return minutes_between(self.started_at, self.completed_at)


@dataclass
class JobInfo:
    id: int
    name: str
    status: str
    conclusion: str
    created_at: dt.datetime | None
    started_at: dt.datetime | None
    completed_at: dt.datetime | None
    labels: list[str] = field(default_factory=list)
    steps: list[StepInfo] = field(default_factory=list)

    @property
    def duration_min(self) -> float | None:
        return minutes_between(self.started_at, self.completed_at)

    @property
    def queue_min(self) -> float | None:
        return minutes_between(self.created_at, self.started_at)


@dataclass
class WorkflowRunInfo:
    id: int
    name: str
    workflow_id: int | None
    status: str
    conclusion: str
    event: str
    html_url: str
    head_sha: str
    created_at: dt.datetime | None
    run_started_at: dt.datetime | None
    updated_at: dt.datetime | None
    schedule_interval_min: float | None = None
    jobs: list[JobInfo] = field(default_factory=list)

    @property
    def duration_min(self) -> float | None:
        start = self.run_started_at or self.created_at
        return minutes_between(start, self.updated_at)

    @property
    def completed_at(self) -> dt.datetime | None:
        return self.updated_at if self.status == "completed" else None


@dataclass
class PullRequestInfo:
    number: int
    title: str
    html_url: str
    author: str
    state: str
    created_at: dt.datetime | None
    merged_at: dt.datetime | None
    head_sha: str
    base_ref: str
    head_ref: str
    workflows: list[WorkflowRunInfo] = field(default_factory=list)

    @property
    def ci_completed_at(self) -> dt.datetime | None:
        completed = [run.completed_at for run in self.workflows if run.completed_at]
        return max(completed) if completed else None

    @property
    def e2e_min(self) -> float | None:
        return minutes_between(self.created_at, self.merged_at)

    @property
    def review_after_ci_min(self) -> float | None:
        return minutes_between(self.ci_completed_at, self.merged_at, allow_negative=True)


class GitHubClient:
    def __init__(self, token: str, sleep: float = 0.0) -> None:
        self.token = token
        self.sleep = sleep
        self.calls = 0
        self.errors = 0
        self.warnings = 0
        self.rate_limit_remaining: int | None = None
        self.rate_limit_reset: int | None = None

    def _handle_response(self, resp, req, url, attempt):
        self.calls += 1
        # Track rate limit headers
        remaining = resp.headers.get("x-ratelimit-remaining")
        reset = resp.headers.get("x-ratelimit-reset")
        if remaining and remaining.isdigit():
            self.rate_limit_remaining = int(remaining)
        if reset and reset.isdigit():
            self.rate_limit_reset = int(reset)
        
        # Warn if rate limit is getting low
        if self.rate_limit_remaining is not None and self.rate_limit_remaining < 100:
            print(f"WARNING: Rate limit remaining: {self.rate_limit_remaining}, resets at {dt.datetime.fromtimestamp(self.rate_limit_reset, tz=dt.timezone.utc).isoformat()}", file=sys.stderr)
        
        if self.sleep:
            time.sleep(self.sleep)
        headers = {k.lower(): v for k, v in resp.headers.items()}
        return json.loads(resp.read().decode("utf-8")), headers

    def get(self, path: str, params: dict[str, Any] | None = None) -> tuple[Any, dict[str, str]]:
        if params:
            path = f"{path}?{urllib.parse.urlencode(params)}"
        url = path if path.startswith("http") else f"{API_ROOT}{path}"
        req = urllib.request.Request(
            url,
            headers={
                "Accept": "application/vnd.github+json",
                "Authorization": f"Bearer {self.token}",
                "X-GitHub-Api-Version": "2022-11-28",
                "User-Agent": "github-ci-efficiency-report",
            },
        )
        for attempt in range(4):
            try:
                with urllib.request.urlopen(req, timeout=60) as resp:
                    return self._handle_response(resp, req, url, attempt)
            except urllib.error.HTTPError as exc:
                body = exc.read().decode("utf-8", errors="replace")
                if exc.code in {403, 429} and attempt < 3:
                    reset = exc.headers.get("x-ratelimit-reset")
                    delay = 2 ** attempt
                    if reset and reset.isdigit():
                        delay = max(delay, min(60, int(reset) - int(time.time()) + 2))
                    print(f"Rate limited or throttled; sleeping {delay}s before retry", file=sys.stderr)
                    time.sleep(delay)
                    continue
                self.errors += 1
                self.warnings += 1
                print(f"WARNING: GitHub API error {exc.code} for {url}: {body[:200]}", file=sys.stderr)
                raise RuntimeError(f"GitHub API error {exc.code} for {url}: {body}") from exc
            except (urllib.error.URLError, TimeoutError, socket.timeout) as exc:
                if attempt < 3:
                    delay = 2 ** attempt
                    print(f"Temporary network error; sleeping {delay}s before retry: {exc}", file=sys.stderr)
                    time.sleep(delay)
                    continue
                self.errors += 1
                self.warnings += 1
                print(f"WARNING: GitHub API network error for {url}: {exc}", file=sys.stderr)
                raise RuntimeError(f"GitHub API network error for {url}: {exc}") from exc
        self.errors += 1
        self.warnings += 1
        raise RuntimeError(f"GitHub API failed for {url}")

    def paginate(self, path: str, params: dict[str, Any]) -> list[Any]:
        items: list[Any] = []
        next_url: str | None = None
        while True:
            if next_url:
                payload, headers = self.get(next_url)
            else:
                payload, headers = self.get(path, params)
            page_items = extract_page_items(payload)
            items.extend(page_items)
            next_url = parse_next_link(headers.get("link", ""))
            if not next_url:
                break
        return items


def extract_page_items(payload: Any) -> list[Any]:
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, dict):
        return []
    for key in ("items", "workflow_runs", "jobs", "artifacts"):
        value = payload.get(key)
        if isinstance(value, list):
            return value
    return []


def parse_next_link(link_header: str) -> str | None:
    for part in link_header.split(","):
        section = part.strip().split(";")
        if len(section) < 2:
            continue
        url = section[0].strip()[1:-1]
        rel = section[1].strip()
        if rel == 'rel="next"':
            return url
    return None


def get_merged_pr_numbers(client: GitHubClient, repo: str, since: str, until: str, max_prs: int | None) -> list[int]:
    query = f"repo:{repo} is:pr is:merged merged:{since}..{until}"
    params = {"q": query, "sort": "updated", "order": "desc", "per_page": 100}
    payload, _ = client.get("/search/issues", params)
    total = payload.get("total_count", 0)
    items = payload.get("items", [])
    if total > 1000:
        print("Warning: GitHub Search returns at most 1000 results. Use a smaller date range.", file=sys.stderr)
    page = 2
    while len(items) < min(total, 1000) and (max_prs is None or len(items) < max_prs):
        params["page"] = page
        page_payload, _ = client.get("/search/issues", params)
        page_items = page_payload.get("items", [])
        if not page_items:
            break
        items.extend(page_items)
        page += 1
    numbers = [int(item["number"]) for item in items]
    return numbers[:max_prs] if max_prs else numbers


def fetch_pr(client: GitHubClient, repo: str, number: int) -> PullRequestInfo | None:
    data, _ = client.get(f"/repos/{repo}/pulls/{number}")
    if not data.get("merged_at"):
        return None
    return PullRequestInfo(
        number=number,
        title=data.get("title", ""),
        html_url=data.get("html_url", ""),
        author=(data.get("user") or {}).get("login", ""),
        state=data.get("state", ""),
        created_at=parse_dt(data.get("created_at")),
        merged_at=parse_dt(data.get("merged_at")),
        head_sha=(data.get("head") or {}).get("sha", ""),
        base_ref=(data.get("base") or {}).get("ref", ""),
        head_ref=(data.get("head") or {}).get("ref", ""),
    )


def fetch_runs_for_sha(client: GitHubClient, repo: str, sha: str) -> list[WorkflowRunInfo]:
    runs = client.paginate(
        f"/repos/{repo}/actions/runs",
        {"head_sha": sha, "per_page": 100, "exclude_pull_requests": "false"},
    )
    result = []
    for run in runs:
        result.append(
            WorkflowRunInfo(
                id=int(run["id"]),
                name=run.get("name") or str(run.get("workflow_id") or ""),
                workflow_id=run.get("workflow_id"),
                status=run.get("status", ""),
                conclusion=run.get("conclusion") or "",
                event=run.get("event", ""),
                html_url=run.get("html_url", ""),
                head_sha=run.get("head_sha", ""),
                created_at=parse_dt(run.get("created_at")),
                run_started_at=parse_dt(run.get("run_started_at")),
                updated_at=parse_dt(run.get("updated_at")),
            )
        )
    return result


def fetch_jobs_for_run(client: GitHubClient, repo: str, run_id: int) -> list[JobInfo]:
    jobs = client.paginate(f"/repos/{repo}/actions/runs/{run_id}/jobs", {"per_page": 100})
    result = []
    for job in jobs:
        steps = [
            StepInfo(
                name=step.get("name", ""),
                status=step.get("status", ""),
                conclusion=step.get("conclusion") or "",
                number=step.get("number"),
                started_at=parse_dt(step.get("started_at")),
                completed_at=parse_dt(step.get("completed_at")),
            )
            for step in job.get("steps", [])
        ]
        result.append(
            JobInfo(
                id=int(job["id"]),
                name=job.get("name", ""),
                status=job.get("status", ""),
                conclusion=job.get("conclusion") or "",
                created_at=parse_dt(job.get("created_at")),
                started_at=parse_dt(job.get("started_at")),
                completed_at=parse_dt(job.get("completed_at")),
                labels=job.get("labels", []),
                steps=steps,
            )
        )
    return result


def fetch_schedule_interval(client: GitHubClient, repo: str, workflow_id: int) -> float | None:
    """Fetch a workflow definition and extract the minimum schedule interval in minutes."""
    try:
        data, _ = client.get(f"/repos/{repo}/actions/workflows/{workflow_id}")
    except RuntimeError:
        return None
    if not isinstance(data, dict):
        return None
    workflow = data.get("workflow", data)
    if not isinstance(workflow, dict):
        return None
    path = workflow.get("path", "")
    if not path:
        return None
    try:
        file_data, _ = client.get(f"/repos/{repo}/contents/{path}")
    except RuntimeError:
        return None
    if not isinstance(file_data, dict):
        return None
    import base64
    content_b64 = file_data.get("content", "")
    try:
        content = base64.b64decode(content_b64).decode("utf-8")
    except Exception:
        return None
    cron_exprs = []
    for line in content.splitlines():
        stripped = line.strip()
        if stripped.startswith("- cron:") or stripped.startswith("cron:"):
            cron_val = stripped.split("cron:", 1)[1].strip().strip("'\"")
            if cron_val:
                cron_exprs.append(cron_val)
    return parse_schedule_interval(cron_exprs)


def estimate_api_calls(pr_count: int, runs_per_pr_avg: float = 2.0, jobs_per_run_avg: float = 3.0, schedule_ratio: float = 0.1) -> dict[str, Any]:
    """Estimate total API calls needed for the report.
    
    Returns a dict with estimated calls breakdown and warnings.
    """
    # Search PRs (may need pagination)
    search_calls = max(1, min(pr_count // 100 + 1, 10))
    
    # Fetch each PR details
    pr_fetch_calls = pr_count
    
    # Fetch runs for each PR (paginated, avg 2 runs per PR)
    total_runs = int(pr_count * runs_per_pr_avg)
    run_fetch_calls = max(total_runs, pr_count)  # At least 1 call per PR
    
    # Fetch jobs for each run (paginated, avg 3 jobs per run)
    job_fetch_calls = total_runs
    
    # Fetch schedule intervals (for scheduled workflows)
    schedule_workflows = int(total_runs * schedule_ratio)
    schedule_calls = schedule_workflows * 2  # 2 calls per workflow (definition + content)
    
    total = search_calls + pr_fetch_calls + run_fetch_calls + job_fetch_calls + schedule_calls
    
    # GitHub API limits: 5000/hour for authenticated users
    # Search API: 30/minute
    api_limit = 5000
    search_limit = 30  # per minute
    
    warnings = []
    if total > api_limit:
        warnings.append(f"Estimated {total} API calls exceeds hourly limit of {api_limit}. Use --max-prs or a smaller date range.")
    if search_calls > search_limit:
        warnings.append(f"Estimated {search_calls} search calls may exceed 30/minute search rate limit.")
    if total > api_limit * 0.8:
        warnings.append(f"Estimated {total} calls is >80% of hourly limit ({api_limit}). Consider splitting the date range.")
    
    return {
        "total": total,
        "search": search_calls,
        "pr_fetch": pr_fetch_calls,
        "run_fetch": run_fetch_calls,
        "job_fetch": job_fetch_calls,
        "schedule": schedule_calls,
        "api_limit": api_limit,
        "warnings": warnings,
    }


def collect_report(client: GitHubClient, repo: str, since: str, until: str, max_prs: int | None, concurrency: int = 5) -> list[PullRequestInfo]:
    pr_numbers = get_merged_pr_numbers(client, repo, since, until, max_prs)
    if not pr_numbers:
        print("No merged PRs found in the given date range.", file=sys.stderr)
        return []
    
    prs: list[PullRequestInfo] = []
    run_cache: dict[str, list[WorkflowRunInfo]] = {}
    job_cache: dict[int, list[JobInfo]] = {}
    schedule_cache: dict[int, float | None] = {}
    failed_prs: list[int] = []
    failed_runs: list[int] = []
    
    # Phase 1: Fetch PR details concurrently
    print(f"Fetching {len(pr_numbers)} PR details with concurrency={concurrency}...", file=sys.stderr)
    pr_results: dict[int, PullRequestInfo | None] = {}
    
    def fetch_one_pr(number: int) -> tuple[int, PullRequestInfo | None]:
        try:
            pr = fetch_pr(client, repo, number)
            return number, pr
        except Exception as e:
            return number, None
    
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {executor.submit(fetch_one_pr, num): num for num in pr_numbers}
        for future in as_completed(futures):
            number, pr = future.result()
            if pr:
                pr_results[number] = pr
            else:
                failed_prs.append(number)
    
    if failed_prs:
        print(f"WARNING: Failed to fetch {len(failed_prs)} PRs: #{', #'.join(map(str, failed_prs[:10]))}{'...' if len(failed_prs) > 10 else ''}", file=sys.stderr)
    
    # Collect unique SHAs for batch run fetching
    sha_to_prs: dict[str, list[PullRequestInfo]] = defaultdict(list)
    for pr in pr_results.values():
        if pr:
            sha_to_prs[pr.head_sha].append(pr)
    
    # Phase 2: Fetch runs for each unique SHA concurrently
    print(f"Fetching workflow runs for {len(sha_to_prs)} unique SHAs...", file=sys.stderr)
    
    def fetch_runs_for_one_sha(sha: str) -> tuple[str, list[WorkflowRunInfo]]:
        try:
            runs = fetch_runs_for_sha(client, repo, sha)
            return sha, runs
        except Exception as e:
            print(f"WARNING: Failed to fetch runs for SHA {sha[:8]}...: {e}", file=sys.stderr)
            return sha, []
    
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {executor.submit(fetch_runs_for_one_sha, sha): sha for sha in sha_to_prs}
        for future in as_completed(futures):
            sha, runs = future.result()
            run_cache[sha] = runs
    
    # Collect unique run IDs for batch job fetching
    unique_run_ids: set[int] = set()
    for runs in run_cache.values():
        for run in runs:
            unique_run_ids.add(run.id)
    
    # Phase 3: Fetch jobs for each run concurrently
    print(f"Fetching jobs for {len(unique_run_ids)} workflow runs...", file=sys.stderr)
    
    def fetch_jobs_for_one_run(run_id: int) -> tuple[int, list[JobInfo]]:
        try:
            jobs = fetch_jobs_for_run(client, repo, run_id)
            return run_id, jobs
        except Exception as e:
            print(f"WARNING: Failed to fetch jobs for run {run_id}: {e}", file=sys.stderr)
            failed_runs.append(run_id)
            return run_id, []
    
    with ThreadPoolExecutor(max_workers=concurrency) as executor:
        futures = {executor.submit(fetch_jobs_for_one_run, run_id): run_id for run_id in unique_run_ids}
        for future in as_completed(futures):
            run_id, jobs = future.result()
            job_cache[run_id] = jobs
    
    if failed_runs:
        print(f"WARNING: Failed to fetch jobs for {len(failed_runs)} runs", file=sys.stderr)
    
    # Phase 4: Fetch schedule intervals for scheduled workflows
    schedule_workflows = set()
    for runs in run_cache.values():
        for run in runs:
            if run.event == "schedule" and run.workflow_id is not None:
                schedule_workflows.add(run.workflow_id)
    
    if schedule_workflows:
        print(f"Fetching schedule intervals for {len(schedule_workflows)} workflows...", file=sys.stderr)
        
        def fetch_schedule_for_one_workflow(wf_id: int) -> tuple[int, float | None]:
            try:
                interval = fetch_schedule_interval(client, repo, wf_id)
                return wf_id, interval
            except Exception as e:
                print(f"WARNING: Failed to fetch schedule for workflow {wf_id}: {e}", file=sys.stderr)
                return wf_id, None
        
        with ThreadPoolExecutor(max_workers=concurrency) as executor:
            futures = {executor.submit(fetch_schedule_for_one_workflow, wf_id): wf_id for wf_id in schedule_workflows}
            for future in as_completed(futures):
                wf_id, interval = future.result()
                schedule_cache[wf_id] = interval
    
    # Assemble results
    for number in pr_numbers:
        pr = pr_results.get(number)
        if not pr:
            continue
        if pr.head_sha in run_cache:
            pr.workflows = run_cache[pr.head_sha]
        for run in pr.workflows:
            if run.id in job_cache:
                run.jobs = job_cache[run.id]
            if run.event == "schedule" and run.workflow_id is not None:
                run.schedule_interval_min = schedule_cache.get(run.workflow_id)
        prs.append(pr)
    
    # Summary
    print(f"\nCollection summary:", file=sys.stderr)
    print(f"  PRs fetched: {len(prs)}/{len(pr_numbers)} ({len(failed_prs)} failed)", file=sys.stderr)
    print(f"  Runs fetched: {len(run_cache)} unique SHAs", file=sys.stderr)
    print(f"  Jobs fetched: {len(job_cache)} runs", file=sys.stderr)
    print(f"  API calls: {client.calls}", file=sys.stderr)
    if client.warnings > 0:
        print(f"  Warnings: {client.warnings}", file=sys.stderr)
    
    return prs


def write_sheet(ws, rows: list[list[Any]]) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    for row in rows:
        ws.append(row)
    if rows:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 60)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width


def build_workbook(prs: list[PullRequestInfo], repo: str, since: str, until: str, output: str, api_calls: int, step_types_file: str | None = None) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Missing dependency: openpyxl. Install it with: python3 -m pip install openpyxl", file=sys.stderr)
        raise SystemExit(2)

    wb = Workbook()
    wb.remove(wb.active)

    workflow_values: dict[tuple[str, str], list[float]] = defaultdict(list)
    workflow_schedule: dict[tuple[str, str], float | None] = {}
    workflow_queue_values: dict[tuple[str, str], list[float]] = defaultdict(list)
    job_duration_values: dict[str, list[float]] = defaultdict(list)
    job_queue_values: dict[str, list[float]] = defaultdict(list)
    job_resource_type: dict[str, str] = {}
    for pr in prs:
        for run in pr.workflows:
            if run.conclusion == "success" and run.duration_min is not None:
                trigger_type = get_trigger_type(run.event)
                key = (run.name, trigger_type)
                workflow_values[key].append(run.duration_min)
                if run.schedule_interval_min is not None:
                    workflow_schedule[key] = run.schedule_interval_min
                job_queues_in_run = [j.queue_min for j in run.jobs if j.conclusion == "success" and j.queue_min is not None]
                if job_queues_in_run:
                    workflow_queue_values[key].append(max(job_queues_in_run))
            for job in run.jobs:
                if job.conclusion != "success":
                    continue
                key = f"{run.name} / {job.name}"
                if job.duration_min is not None:
                    job_duration_values[key].append(job.duration_min)
                if job.queue_min is not None:
                    job_queue_values[key].append(job.queue_min)
                if key not in job_resource_type and job.labels:
                    job_resource_type[key] = _extract_resource_type(job.labels)

    ws = wb.create_sheet("workflow_stats")
    rows = [["工作流", "触发类型", "运行次数", "平均E2E(分钟)", "P50 E2E(分钟)", "P90 E2E(分钟)", "平均排队(分钟)", "P50 排队(分钟)", "P90 排队(分钟)", "调度周期(分钟)"]]
    for name, trigger_type in sorted(workflow_values):
        values = workflow_values[(name, trigger_type)]
        schedule = workflow_schedule.get((name, trigger_type))
        queues = workflow_queue_values.get((name, trigger_type), [])
        rows.append([name, trigger_type, len(values), average(values), percentile(values, 0.50), percentile(values, 0.90), average(queues), percentile(queues, 0.50), percentile(queues, 0.90), schedule])
    write_sheet(ws, rows)

    ws = wb.create_sheet("job_stats")
    rows = [[
        "工作流/任务",
        "资源类型",
        "执行次数",
        "平均E2E(分钟)",
        "P50 E2E(分钟)",
        "P90 E2E(分钟)",
        "平均排队(分钟)",
        "P50 排队(分钟)",
        "P90 排队(分钟)",
    ]]
    for key in sorted(set(job_duration_values) | set(job_queue_values)):
        durations = job_duration_values[key]
        queues = job_queue_values[key]
        resource = job_resource_type.get(key, "")
        rows.append([
            key,
            resource,
            max(len(durations), len(queues)),
            average(durations),
            percentile(durations, 0.50),
            percentile(durations, 0.90),
            average(queues),
            percentile(queues, 0.50),
            percentile(queues, 0.90),
        ])
    write_sheet(ws, rows)

    # --- step_stats sheet ---
    step_durations: dict[str, list[float]] = defaultdict(list)
    step_conclusions: dict[str, list[str]] = defaultdict(list)
    for pr in prs:
        for run in pr.workflows:
            for job in run.jobs:
                for step in job.steps:
                    if step.name and step.duration_min is not None and step.duration_min >= 3:
                        step_durations[step.name].append(step.duration_min)
                        step_conclusions[step.name].append(step.conclusion)
    all_unique_step_names = sorted(step_durations.keys())

    step_type_map: dict[str, str] = {}
    if step_types_file:
        step_type_map = load_step_types(step_types_file)

    ws = wb.create_sheet("step_stats")
    rows = [[
        "步骤名称",
        "步骤类型",
        "执行次数",
        "平均耗时(分钟)",
        "P50 耗时(分钟)",
        "P90 耗时(分钟)",
        "成功率",
    ]]
    for step_name in all_unique_step_names:
        step_type = step_type_map.get(step_name, "")
        if step_type == "排除":
            continue
        durations = step_durations[step_name]
        conclusions = step_conclusions[step_name]
        success_count = sum(1 for c in conclusions if c == "success")
        success_rate = round(success_count / len(conclusions), 3) if conclusions else None
        rows.append([
            step_name,
            step_type,
            len(durations),
            average(durations),
            percentile(durations, 0.50),
            percentile(durations, 0.90),
            success_rate,
        ])
    write_sheet(ws, rows)

    ws = wb.create_sheet("pr_stats")
    rows = [[
        "PR编号",
        "标题",
        "作者",
        "创建时间",
        "合并时间",
        "CI完成时间",
        "PR E2E(分钟)",
        "CI后评审(分钟)",
        "工作流数量",
        "链接",
    ]]
    for pr in sorted(prs, key=lambda p: p.number):
        rows.append([
            pr.number,
            pr.title,
            pr.author,
            fmt_dt(pr.created_at),
            fmt_dt(pr.merged_at),
            fmt_dt(pr.ci_completed_at),
            pr.e2e_min,
            pr.review_after_ci_min,
            len(pr.workflows),
            pr.html_url,
        ])
    rows.append([])
    e2e = [pr.e2e_min for pr in prs if pr.e2e_min is not None]
    review = [pr.review_after_ci_min for pr in prs if pr.review_after_ci_min is not None]
    rows.append(["汇总", f"仓库={repo}", f"周期={since}..{until}", f"PR数={len(prs)}", f"API调用={api_calls}"])
    rows.append(["指标", "平均(分钟)", "P50(分钟)", "P90(分钟)"])
    rows.append(["PR E2E", average(e2e), percentile(e2e, 0.50), percentile(e2e, 0.90)])
    rows.append(["CI后评审", average(review), percentile(review, 0.50), percentile(review, 0.90)])
    write_sheet(ws, rows)

    ws = wb.create_sheet("pr_details")
    header = [
        "层级",
        "PR编号",
        "PR标题",
        "PR作者",
        "PR创建时间",
        "PR合并时间",
        "PR E2E(分钟)",
        "CI后评审(分钟)",
        "工作流名称",
        "工作流运行ID",
        "工作流状态",
        "工作流结论",
        "工作流创建时间",
        "工作流开始时间",
        "工作流完成时间",
        "工作流E2E(分钟)",
        "任务名称",
        "任务ID",
        "任务状态",
        "任务结论",
        "任务创建时间",
        "任务开始时间",
        "任务完成时间",
        "任务排队(分钟)",
        "任务E2E(分钟)",
        "步骤序号",
        "步骤名称",
        "步骤状态",
        "步骤结论",
        "步骤开始时间",
        "步骤完成时间",
        "步骤E2E(分钟)",
        "链接",
    ]
    ws.append(header)
    for pr in sorted(prs, key=lambda p: p.number):
        ws.append([
            "PR",
            pr.number,
            pr.title,
            pr.author,
            fmt_dt(pr.created_at),
            fmt_dt(pr.merged_at),
            pr.e2e_min,
            pr.review_after_ci_min,
            "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
            pr.html_url,
        ])
        for run in pr.workflows:
            ws.append([
                "WORKFLOW",
                pr.number,
                pr.title,
                pr.author,
                fmt_dt(pr.created_at),
                fmt_dt(pr.merged_at),
                pr.e2e_min,
                pr.review_after_ci_min,
                run.name,
                run.id,
                run.status,
                run.conclusion,
                fmt_dt(run.created_at),
                fmt_dt(run.run_started_at),
                fmt_dt(run.completed_at),
                run.duration_min,
                "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "",
                run.html_url,
            ])
            ws.row_dimensions[ws.max_row].outlineLevel = 1
            for job in run.jobs:
                ws.append([
                    "JOB",
                    pr.number,
                    pr.title,
                    pr.author,
                    fmt_dt(pr.created_at),
                    fmt_dt(pr.merged_at),
                    pr.e2e_min,
                    pr.review_after_ci_min,
                    run.name,
                    run.id,
                    run.status,
                    run.conclusion,
                    fmt_dt(run.created_at),
                    fmt_dt(run.run_started_at),
                    fmt_dt(run.completed_at),
                    run.duration_min,
                    job.name,
                    job.id,
                    job.status,
                    job.conclusion,
                    fmt_dt(job.created_at),
                    fmt_dt(job.started_at),
                    fmt_dt(job.completed_at),
                    job.queue_min,
                    job.duration_min,
                    "", "", "", "", "", "", "",
                    run.html_url,
                ])
                ws.row_dimensions[ws.max_row].outlineLevel = 2
                for step in job.steps:
                    ws.append([
                        "STEP",
                        pr.number,
                        pr.title,
                        pr.author,
                        fmt_dt(pr.created_at),
                        fmt_dt(pr.merged_at),
                        pr.e2e_min,
                        pr.review_after_ci_min,
                        run.name,
                        run.id,
                        run.status,
                        run.conclusion,
                        fmt_dt(run.created_at),
                        fmt_dt(run.run_started_at),
                        fmt_dt(run.completed_at),
                        run.duration_min,
                        job.name,
                        job.id,
                        job.status,
                        job.conclusion,
                        fmt_dt(job.created_at),
                        fmt_dt(job.started_at),
                        fmt_dt(job.completed_at),
                        job.queue_min,
                        job.duration_min,
                        step.number,
                        step.name,
                        step.status,
                        step.conclusion,
                        fmt_dt(step.started_at),
                        fmt_dt(step.completed_at),
                        step.duration_min,
                        run.html_url,
                    ])
                    ws.row_dimensions[ws.max_row].outlineLevel = 3
    write_sheet(ws, [])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in col) + 2, 60)
        ws.column_dimensions[get_column_letter(col[0].column)].width = width

    wb.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export GitHub CI efficiency metrics to Excel.")
    parser.add_argument("--repo", required=True, help="Repository in owner/name form.")
    parser.add_argument("--since", required=True, help="Start date, YYYY-MM-DD, inclusive.")
    parser.add_argument("--until", required=True, help="End date, YYYY-MM-DD, inclusive.")
    parser.add_argument("--output", help="Output .xlsx path.")
    parser.add_argument("--token", default=os.getenv("GITHUB_TOKEN") or os.getenv("GH_TOKEN"), help="GitHub token. Defaults to GITHUB_TOKEN or GH_TOKEN.")
    parser.add_argument("--max-prs", type=int, default=None, help="Optional cap for testing or very large repos.")
    parser.add_argument("--sleep", type=float, default=0.0, help="Optional sleep seconds between API calls.")
    parser.add_argument("--concurrency", type=int, default=5, help="Number of concurrent API requests (default: 5).")
    parser.add_argument("--estimate-only", action="store_true", help="Estimate API calls and exit without running the report.")
    parser.add_argument("--export-step-names", help="Export unique step names to JSON file and exit (for external LLM classification).")
    parser.add_argument("--step-types", help="JSON file mapping step names to types (output of --export-step-names after classification).")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.token:
        print("Provide a GitHub token with --token or GITHUB_TOKEN/GH_TOKEN.", file=sys.stderr)
        return 2
    if "/" not in args.repo:
        print("--repo must be owner/name", file=sys.stderr)
        return 2
    if not args.output and not args.export_step_names and not args.estimate_only:
        print("Provide --output or --export-step-names or --estimate-only", file=sys.stderr)
        return 2
    
    client = GitHubClient(args.token, sleep=args.sleep)
    
    # Phase 0: Estimate API calls
    print(f"Estimating API calls for {args.repo} ({args.since} to {args.until})...", file=sys.stderr)
    estimate = estimate_api_calls(args.max_prs or 100)  # Use 100 as default estimate
    print(f"Estimated API calls: {estimate['total']} (limit: {estimate['api_limit']}/hour)", file=sys.stderr)
    print(f"  Search: {estimate['search']}", file=sys.stderr)
    print(f"  PR fetch: {estimate['pr_fetch']}", file=sys.stderr)
    print(f"  Run fetch: {estimate['run_fetch']}", file=sys.stderr)
    print(f"  Job fetch: {estimate['job_fetch']}", file=sys.stderr)
    print(f"  Schedule fetch: {estimate['schedule']}", file=sys.stderr)
    
    if estimate["warnings"]:
        for warning in estimate["warnings"]:
            print(f"WARNING: {warning}", file=sys.stderr)
    
    if args.estimate_only:
        return 0
    
    # Collect report with concurrency
    start_time = time.time()
    prs = collect_report(client, args.repo, args.since, args.until, args.max_prs, concurrency=args.concurrency)
    elapsed = time.time() - start_time
    
    if not prs:
        print("No PRs to report.", file=sys.stderr)
        return 1

    if args.export_step_names:
        export_step_names(prs, args.export_step_names)
        return 0

    build_workbook(prs, args.repo, args.since, args.until, args.output, client.calls, step_types_file=args.step_types)
    
    # Final summary
    print(f"\n{'='*60}", file=sys.stderr)
    print(f"Report generated: {args.output}", file=sys.stderr)
    print(f"  Merged PRs: {len(prs)}", file=sys.stderr)
    print(f"  API calls: {client.calls}", file=sys.stderr)
    print(f"  Elapsed time: {elapsed:.1f}s", file=sys.stderr)
    print(f"  Requests/sec: {client.calls/elapsed:.2f}" if elapsed > 0 else "", file=sys.stderr)
    if client.warnings > 0:
        print(f"  Warnings: {client.warnings}", file=sys.stderr)
    if client.errors > 0:
        print(f"  Errors: {client.errors}", file=sys.stderr)
    print(f"{'='*60}", file=sys.stderr)
    
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
