#!/usr/bin/env python3
"""CI 耗时分析（目的2）：只统计"成功且跑了所有 job"的 CI run。

过滤标准（ADR-003）：conclusion == "success" 且 run 耗时 > --min-duration（默认 10min）。
效率：先拉 run 列表 → 过滤 → 只对命中 run 抓 jobs(含 steps)，省 API 调用。
输出：HTML（统计 + 自动文本洞察）+ Excel（原始明细）。

用法:
  export GITHUB_TOKEN=$(gh auth token)
  python3 ci_duration_analysis.py \\
    --repo vllm-project/vllm-ascend --workflow E2E \\
    --from 2026-06-24 --to 2026-07-01 \\
    --output-dir reports

数据源：GitHub REST API（当前唯一实现路径）。Turso DB 为 TODO（ADR-003 备选方案1），
凭证就绪后可在 SQL 层预过滤实现 0 调用，分析/输出逻辑不变。
"""
from __future__ import annotations

import argparse
import html as html_lib
import json
import os
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from pathlib import Path

# 复用现有 by-SHA adapter 的客户端/数据类/工具函数（一处实现，多处复用）
SCRIPT_DIR = Path(__file__).parent
SKILL_SCRIPTS = SCRIPT_DIR / "skills" / "github-ci-efficiency-report" / "scripts"
sys.path.insert(0, str(SKILL_SCRIPTS))
sys.path.insert(0, str(SCRIPT_DIR))  # workflow_runs_on_date.py 在上层目录
from github_ci_efficiency_report import (  # noqa: E402
    GitHubClient,
    WorkflowRunInfo,
    JobInfo,
    parse_dt,
    percentile,
    average,
    write_sheet,
)
from workflow_runs_on_date import fetch_workflow_id, norm_job_name  # noqa: E402

DEFAULT_STEP_NAMES = SCRIPT_DIR / "step-names.json"
DEFAULT_MIN_DURATION = 10.0  # 分钟；E2E 完整路径 ~50min+，lint 路径 ≤5min

# ponytail: step-names.json 是静态映射，会漏新 step。加一个正则兜底分类器，
# 让洞察里的 step 类型占比不失真（如 'Run selected tests' 没映射时不会全归到"其他"）。
# 升级路径：step-names.json 补全后此兜底自然失效。
_STEP_RE_BUILD = re.compile(r"install|checkout|cache|setup|compile|build|rebase|restore|get\s+(csrc|arch)|csrc", re.I)
_STEP_RE_TEST = re.compile(r"run\s+.*test|pre-commit|mypy|linkcheck|flake8|ruff|e2e", re.I)
_STEP_RE_EXCLUDE = re.compile(r"post\s|stop\s+(container|runner)|complete\s+job|clean\s+up", re.I)


def classify_step(name: str, step_map: dict[str, str]) -> str:
    """优先查静态映射；未命中则正则兜底。"""
    t = step_map.get(name)
    if t:
        return t
    if _STEP_RE_EXCLUDE.search(name):
        return "排除"
    if _STEP_RE_TEST.search(name):
        return "执行测试"
    if _STEP_RE_BUILD.search(name):
        return "构建"
    return "CI启动"  # 默认归 CI启动（Set up job/Initialize containers 等）


# ─── 数据获取 ──────────────────────────────────────────────────────────

def fetch_runs_in_range(client: GitHubClient, repo: str, workflow_id: int,
                        date_from: str, date_to: str) -> list[WorkflowRunInfo]:
    """按 created 日期范围拉某 workflow 的所有 run。

    GitHub Actions runs 列表 API 有 1000 条硬上限（~10 页×100），超过会静默截断。
    因此按天拆分查询再合并，每天远不到 1000。 ponytail: 上限是 GitHub 侧的，
    升级路径是切 Turso（ADR-001，0 调用无上限）。
    """
    seen_ids: set[int] = set()
    all_runs: list[WorkflowRunInfo] = []
    d0 = datetime.strptime(date_from, "%Y-%m-%d").date()
    d1 = datetime.strptime(date_to, "%Y-%m-%d").date()
    day = timedelta(days=1)
    cur = d0
    while cur <= d1:
        ds = cur.isoformat()
        runs = _fetch_runs_one_day(client, repo, workflow_id, ds)
        for r in runs:
            if r.id not in seen_ids:
                seen_ids.add(r.id)
                all_runs.append(r)
        cur += day
    # 按创建时间倒序，与原行为一致
    all_runs.sort(key=lambda x: x.created_at or parse_dt("2000-01-01"), reverse=True)
    return all_runs


def _fetch_runs_one_day(client: GitHubClient, repo: str, workflow_id: int, date: str) -> list[WorkflowRunInfo]:
    """拉某 workflow 在指定日期(UTC)创建的所有 run。"""
    runs: list[WorkflowRunInfo] = []
    params = {"created": date, "per_page": 100}
    page = 1
    while True:
        p = dict(params)
        if page > 1:
            p["page"] = page
        data, headers = client.get(f"/repos/{repo}/actions/workflows/{workflow_id}/runs", p)
        items = data.get("workflow_runs", [])
        if not items:
            break
        for run in items:
            runs.append(WorkflowRunInfo(
                id=int(run["id"]),
                name=run.get("name") or str(run.get("workflow_id") or ""),
                workflow_id=run.get("workflow_id"),
                status=run.get("status", ""),
                conclusion=run.get("conclusion") or "",
                event=run.get("event", ""),
                html_url=run.get("html_url", ""),
                head_sha=run.get("head_sha", ""),
                created_at=parse_dt(run.get("created_at")),
                run_started_at=parse_dt(run.get("run_started_at")),
                updated_at=parse_dt(run.get("updated_at")),
            ))
        page += 1
        if 'rel="next"' not in headers.get("link", ""):
            break
    return runs


def filter_runs(runs: list[WorkflowRunInfo], min_duration: float) -> list[WorkflowRunInfo]:
    """目的2 过滤：成功 + 耗时 > min_duration。"""
    out = []
    for r in runs:
        if r.conclusion != "success":
            continue
        d = r.duration_min
        if d is None or d <= min_duration:
            continue
        out.append(r)
    return out


def fetch_jobs_for_runs(client: GitHubClient, repo: str, run_ids: list[int],
                        concurrency: int) -> dict[int, list[JobInfo]]:
    """并发抓 jobs(含 steps)，只对过滤后的 run 调 API。复用 fetch_jobs_for_run。"""
    from github_ci_efficiency_report import fetch_jobs_for_run
    run_jobs: dict[int, list[JobInfo]] = {}
    failed: list[int] = []

    def one(rid: int):
        try:
            return rid, fetch_jobs_for_run(client, repo, rid)
        except Exception as e:  # noqa: BLE001
            print(f"WARNING: run {rid} jobs 失败: {e}", file=sys.stderr)
            return rid, []

    with ThreadPoolExecutor(max_workers=concurrency) as ex:
        futs = {ex.submit(one, rid): rid for rid in run_ids}
        for fut in as_completed(futs):
            rid, jobs = fut.result()
            run_jobs[rid] = jobs
            if not jobs:
                failed.append(rid)
    if failed:
        print(f"WARNING: {len(failed)} 个 run 无 job 数据", file=sys.stderr)
    return run_jobs


# ─── 统计 ──────────────────────────────────────────────────────────────

def _rows_job_stats(run_jobs: dict[int, list[JobInfo]]) -> list[dict]:
    groups: dict[str, dict] = defaultdict(lambda: {"dur": [], "queue": [], "succ": 0, "total": 0})
    for jobs in run_jobs.values():
        for j in jobs:
            if j.conclusion != "success":  # 目的2：只统计成功 job 的耗时
                continue
            key = norm_job_name(j.name)
            if j.duration_min is not None:
                groups[key]["dur"].append(j.duration_min)
            if j.queue_min is not None:
                groups[key]["queue"].append(j.queue_min)
            groups[key]["total"] += 1
            groups[key]["succ"] += 1
    rows = []
    for key, g in sorted(groups.items(), key=lambda x: -sum(x[1]["dur"])):
        d, q = g["dur"], g["queue"]
        rows.append({
            "任务(归一化)": key, "执行次数": g["total"],
            "平均E2E(分钟)": average(d), "P50(分钟)": percentile(d, 0.5),
            "P90(分钟)": percentile(d, 0.9), "最大(分钟)": max(d) if d else None,
            "平均排队(分钟)": average(q), "P90排队(分钟)": percentile(q, 0.9),
            "总耗时(分钟)": round(sum(d), 1) if d else 0,
        })
    return rows


def _rows_step_stats(run_jobs: dict[int, list[JobInfo]], step_map: dict[str, str]) -> list[dict]:
    dur: dict[str, list[float]] = defaultdict(list)
    for jobs in run_jobs.values():
        for j in jobs:
            if j.conclusion != "success":
                continue
            for s in j.steps:
                if s.duration_min is None or s.duration_min <= 0:
                    continue
                dur[s.name].append(s.duration_min)
    rows = []
    for name in sorted(dur, key=lambda k: -sum(dur[k])):
        if step_map.get(name) == "排除":
            continue
        d = dur[name]
        rows.append({
            "步骤名称": name, "步骤类型": classify_step(name, step_map),
            "执行次数": len(d), "平均耗时(分钟)": average(d),
            "P50(分钟)": percentile(d, 0.5), "P90(分钟)": percentile(d, 0.9),
            "总耗时(分钟)": round(sum(d), 1),
        })
    return rows


def _rows_run_details(runs: list[WorkflowRunInfo]) -> list[dict]:
    rows = []
    for r in sorted(runs, key=lambda x: x.created_at or parse_dt("2000-01-01")):
        rows.append({
            "运行ID": r.id, "创建时间": r.created_at.isoformat(timespec="seconds") if r.created_at else "",
            "事件": r.event, "耗时(分钟)": r.duration_min, "Job数": len(r.jobs),
            "链接": r.html_url,
        })
    return rows


def _rows_job_details(runs: list[WorkflowRunInfo]) -> list[dict]:
    rows = []
    for r in runs:
        for j in sorted(r.jobs, key=lambda x: -(x.duration_min or 0)):
            rows.append({
                "运行ID": r.id, "任务(原始)": j.name, "任务(归一化)": norm_job_name(j.name),
                "任务结论": j.conclusion, "排队(分钟)": j.queue_min, "E2E(分钟)": j.duration_min,
                "步骤数": len(j.steps), "链接": r.html_url,
            })
    return rows


def _rows_step_details(runs: list[WorkflowRunInfo], step_map: dict[str, str]) -> list[dict]:
    rows = []
    for r in runs:
        for j in r.jobs:
            for s in sorted(j.steps, key=lambda x: x.number or 0):
                rows.append({
                    "运行ID": r.id, "任务(归一化)": norm_job_name(j.name),
                    "步骤序号": s.number, "步骤名称": s.name,
                    "步骤类型": classify_step(s.name, step_map), "步骤结论": s.conclusion,
                    "耗时(分钟)": s.duration_min,
                })
    return rows


# ─── 自动文本洞察 ─────────────────────────────────────────────────────

def infer_resource_type(job_name: str) -> str:
    """从 job 名推断硬件/资源类型。复用 workflow_runs_on_date 的归一化后判断。"""
    n = job_name.lower()
    if "a2" in n: return "a2 (2卡)"
    if "a3-4" in n or "a3 4" in n or "a34" in n: return "a3-4 (4卡)"
    if "a3-2" in n or "a3 2" in n or "a32" in n: return "a3-2 (2卡)"
    if "a3" in n: return "a3"
    if "310p-4" in n or "310p 4" in n: return "310p-4"
    if "310p" in n: return "310p-1"
    if "cpu" in n: return "cpu"
    if "lint" in n or "select" in n: return "lint/调度"
    return "其他"


def generate_insights(runs: list[WorkflowRunInfo], run_jobs: dict[int, list[JobInfo]],
                      step_map: dict[str, str]) -> list[str]:
    """数据驱动的自动文本洞察。每条一句话，附数值。"""
    issues = generate_top_issues(runs, run_jobs, step_map)
    if not issues:
        return ["样本不足，无法生成洞察。"]
    return [p["summary"] for p in issues]


def generate_top_issues(runs: list[WorkflowRunInfo], run_jobs: dict[int, list[JobInfo]],
                        step_map: dict[str, str]) -> list[dict]:
    """返回 Top 问题列表，每项 {title, summary, evidence: [str...], severity}。

    severity 越大越靠前。证据是可验证的具体数值/典型 run，供 HTML 汇总区展示。
    """
    issues: list[dict] = []
    n_runs = len(runs) or 1

    # --- 聚合基础数据（一处计算，多处复用）---
    longest_job_counter: dict[str, int] = defaultdict(int)
    job_dur: dict[str, list[float]] = defaultdict(list)
    job_queue: dict[str, list[float]] = defaultdict(list)
    step_dur: dict[str, list[float]] = defaultdict(list)
    worst_queue_run: dict[str, tuple] = {}  # job名 -> (run_id, queue_min)

    for r in runs:
        succ_jobs = [j for j in r.jobs if j.conclusion == "success" and j.duration_min]
        if succ_jobs:
            longest = max(succ_jobs, key=lambda x: x.duration_min)
            longest_job_counter[norm_job_name(longest.name)] += 1
        for j in r.jobs:
            if j.conclusion != "success":
                continue
            key = norm_job_name(j.name)
            if j.duration_min:
                job_dur[key].append(j.duration_min)
            if j.queue_min is not None:
                job_queue[key].append(j.queue_min)
                # 跟踪该 job 类型最差排队的 run 作为典型证据
                if key not in worst_queue_run or j.queue_min > worst_queue_run[key][1]:
                    worst_queue_run[key] = (r.id, j.queue_min)
            for s in j.steps:
                if s.duration_min and s.duration_min > 0:
                    step_dur[s.name].append(s.duration_min)

    # --- 问题1: 排队瓶颈（severity 最高，因为是根因而非症状）---
    queue_issues = []
    for key, qs in job_queue.items():
        if len(qs) < 3:
            continue
        avg_q = average(qs) or 0
        p90_q = percentile(qs, 0.9) or 0
        avg_exec = average(job_dur.get(key, [0])) or 0
        if (avg_exec > 0 and avg_q / avg_exec > 0.5) or p90_q > 30:
            queue_issues.append((key, avg_q, p90_q, avg_exec, len(qs)))
    if queue_issues:
        queue_issues.sort(key=lambda x: -x[2])
        k, aq, p90q, ae, cnt = queue_issues[0]
        ev_run, ev_q = worst_queue_run.get(k, (None, None))
        evidence = [
            f"{k}：平均排队 {aq:.0f}min / P90 排队 {p90q:.0f}min，但平均执行仅 {ae:.0f}min（{cnt} 次）",
            f"{k}：平均执行 {ae:.0f}min，排队占比 {aq/(ae+aq)*100:.0f}%" if (ae+aq) else "",
        ]
        if ev_run:
            evidence.append(f"典型证据：run {ev_run} 中该 job 排队 {ev_q:.0f}min 才开始")
        issues.append({
            "title": "排队瓶颈：硬件池容量不足",
            "summary": (
                f"『{k}』平均排队 {aq:.0f}min、P90 排队 {p90q:.0f}min（执行才 {ae:.0f}min），"
                f"是拉长 run 墙钟的头号根因。"
            ),
            "evidence": [e for e in evidence if e],
            "severity": 100,
        })

    # --- 问题2: 关键路径 job（决定墙钟的最长 job）---
    if longest_job_counter:
        top_cp, top_cp_cnt = max(longest_job_counter.items(), key=lambda x: x[1])
        rate = top_cp_cnt / n_runs * 100
        cp_avg = average(job_dur.get(top_cp, [0])) or 0
        cp_p90 = percentile(job_dur.get(top_cp, [0]), 0.9) or 0
        issues.append({
            "title": f"关键路径：{top_cp}",
            "summary": (
                f"『{top_cp}』在 {top_cp_cnt}/{n_runs} 个 run（{rate:.0f}%）中是最长 job，"
                f"矩阵并行下它决定 run 墙钟。"
            ),
            "evidence": [
                f"{top_cp}：平均执行 {cp_avg:.0f}min / P90 {cp_p90:.0f}min（{len(job_dur.get(top_cp,[]))} 次）",
                f"作为最长 job 出现 {top_cp_cnt} 次，占比 {rate:.0f}%",
            ],
            "severity": 80,
        })

    # --- 问题3: 最耗时 step（执行测试主体）---
    if step_dur:
        top_step, top_durs = max(step_dur.items(), key=lambda x: sum(x[1]))
        # step 类型占比
        type_dur: dict[str, list[float]] = defaultdict(list)
        for jobs in run_jobs.values():
            for j in jobs:
                if j.conclusion != "success":
                    continue
                for s in j.steps:
                    if s.duration_min and s.duration_min > 0:
                        type_dur[classify_step(s.name, step_map)].append(s.duration_min)
        type_totals = [(t, sum(v)) for t, v in type_dur.items() if t != "排除"]
        grand = sum(x[1] for x in type_totals) or 1
        test_pct = next((v / grand * 100 for t, v in type_totals if t == "执行测试"), 0)
        issues.append({
            "title": f"step 热点：{top_step}",
            "summary": (
                f"『{top_step}』总耗时 {sum(top_durs):.0f}min、均 {average(top_durs):.0f}min，"
                f"执行测试类 step 占总耗时 {test_pct:.0f}%。"
            ),
            "evidence": [
                f"{top_step}：执行 {len(top_durs)} 次，均 {average(top_durs):.0f}min，总 {sum(top_durs):.0f}min",
                f"step 类型占比：执行测试 {test_pct:.0f}%、"
                + "、".join(f"{t} {v/grand*100:.0f}%" for t, v in type_totals if t != "执行测试"),
            ],
            "severity": 60,
        })

    issues.sort(key=lambda x: -x["severity"])
    return issues[:3]



# ─── Excel 输出（原始明细）──────────────────────────────────────────────

def write_excel(filepath: str, runs: list[WorkflowRunInfo], run_jobs: dict[int, list[JobInfo]],
                step_map: dict[str, str]):
    try:
        from openpyxl import Workbook
    except ImportError:
        print("⚠ openpyxl 未安装，跳过 Excel 输出（pip install openpyxl）", file=sys.stderr)
        return
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("run_details"); write_sheet(ws, _to_rows(_rows_run_details(runs)))
    ws = wb.create_sheet("job_stats"); write_sheet(ws, _to_rows(_rows_job_stats(run_jobs)))
    ws = wb.create_sheet("job_details"); write_sheet(ws, _to_rows(_rows_job_details(runs)))
    ws = wb.create_sheet("step_stats"); write_sheet(ws, _to_rows(_rows_step_stats(run_jobs, step_map)))
    ws = wb.create_sheet("step_details"); write_sheet(ws, _to_rows(_rows_step_details(runs, step_map)))
    wb.save(filepath)
    print(f"✅ Excel 原始数据: {filepath}", file=sys.stderr)


def _to_rows(dicts: list[dict]) -> list[list]:
    """dict 列表 → write_sheet 需要的 [[header...], [row...]] 格式。"""
    if not dicts:
        return [[]]
    headers = list(dicts[0].keys())
    return [headers] + [[d.get(h) for h in headers] for d in dicts]


# ─── HTML 输出（统计 + 洞察）────────────────────────────────────────────

def _fmt(v, suffix="") -> str:
    if v is None:
        return "-"
    if isinstance(v, float):
        return f"{v:.1f}{suffix}"
    return f"{v}{suffix}"


def _stats_table(title: str, rows: list[dict], note: str = "") -> str:
    if not rows:
        return f"<h3>{html_lib.escape(title)}</h3><p class='muted'>无数据</p>"
    headers = list(rows[0].keys())
    th = "".join(f"<th>{html_lib.escape(str(h))}</th>" for h in headers)
    trs = []
    for r in rows[:50]:  # ponytail: HTML 只展示前 50 行，全量见 Excel
        tds = "".join(f"<td>{html_lib.escape(_fmt(r.get(h)))}</td>" for h in headers)
        trs.append(f"<tr>{tds}</tr>")
    note_html = f"<p class='muted'>{html_lib.escape(note)}</p>" if note else ""
    return f"<h3>{html_lib.escape(title)}</h3>{note_html}<table><thead><tr>{th}</tr></thead><tbody>{''.join(trs)}</tbody></table>"


def write_html(filepath: str, repo: str, workflow: str, date_from: str, date_to: str,
               total_runs: int, filtered_runs: list[WorkflowRunInfo],
               run_jobs: dict[int, list[JobInfo]], step_map: dict[str, str],
               min_duration: float, api_calls: int, elapsed: float):
    top_issues = generate_top_issues(filtered_runs, run_jobs, step_map)
    job_rows = _rows_job_stats(run_jobs)
    step_rows = _rows_step_stats(run_jobs, step_map)
    run_rows = _rows_run_details(filtered_runs)

    # run 级耗时汇总
    run_durs = [r.duration_min for r in filtered_runs if r.duration_min is not None]

    # Top 问题 HTML（每个问题：标题 + 摘要 + 证据列表）
    issue_cards = []
    for idx, issue in enumerate(top_issues, 1):
        sev_cls = f"issue-sev{min(idx, 3)}"
        ev_lis = "".join(f"<li>{html_lib.escape(e)}</li>" for e in issue.get("evidence", []))
        issue_cards.append(f"""
        <div class="issue">
          <div class="issue-head {sev_cls}">#{idx} {html_lib.escape(issue['title'])}</div>
          <div class="issue-body">
            <div class="issue-summary">{html_lib.escape(issue['summary'])}</div>
            <ul class="issue-evidence">{ev_lis}</ul>
          </div>
        </div>""")
    issues_html = f'<h2>🚨 Top 问题（按严重度）</h2><div class="issues">{chr(10).join(issue_cards)}</div>' if issue_cards else ""

    summary_cards = ""
    if run_durs:
        summary_cards = f"""
        <div class="cards">
          <div class="card"><div class="num">{len(filtered_runs)}</div><div class="lab">命中 run 数</div></div>
          <div class="card"><div class="num">{_fmt(average(run_durs))}</div><div class="lab">平均耗时(分钟)</div></div>
          <div class="card"><div class="num">{_fmt(percentile(run_durs, 0.5))}</div><div class="lab">P50(分钟)</div></div>
          <div class="card"><div class="num">{_fmt(percentile(run_durs, 0.9))}</div><div class="lab">P90(分钟)</div></div>
          <div class="card"><div class="num">{_fmt(max(run_durs))}</div><div class="lab">最大(分钟)</div></div>
          <div class="card"><div class="num">{total_runs}</div><div class="lab">总 run 数(过滤前)</div></div>
        </div>"""

    doc = f"""<!DOCTYPE html>
<html lang="zh"><head><meta charset="utf-8"><title>CI 耗时分析 - {html_lib.escape(workflow)}</title>
<style>
  body {{ font-family: -apple-system, "Segoe UI", Roboto, sans-serif; margin: 24px auto; max-width: 1200px; color: #1f2328; }}
  h1 {{ border-bottom: 2px solid #4472C4; padding-bottom: 8px; }}
  h2 {{ color: #4472C4; margin-top: 32px; }}
  h3 {{ margin-top: 24px; }}
  .meta {{ color: #6b7280; font-size: 14px; margin-bottom: 16px; }}
  .cards {{ display: flex; flex-wrap: wrap; gap: 12px; margin: 16px 0; }}
  .card {{ background: #f0f5ff; border-radius: 8px; padding: 12px 20px; min-width: 110px; text-align: center; }}
  .card .num {{ font-size: 24px; font-weight: 700; color: #2c5cc5; }}
  .card .lab {{ font-size: 12px; color: #6b7280; margin-top: 4px; }}
  .insights {{ background: #fff8e1; border-left: 4px solid #f59e0b; padding: 12px 20px; border-radius: 4px; margin: 16px 0; }}
  .insights ul {{ margin: 8px 0; padding-left: 20px; }}
  .insights li {{ margin: 6px 0; line-height: 1.6; }}
  table {{ border-collapse: collapse; width: 100%; margin: 8px 0 24px; font-size: 13px; }}
  th {{ background: #4472C4; color: #fff; padding: 8px 10px; text-align: left; position: sticky; top: 0; }}
  td {{ border: 1px solid #e1e4e8; padding: 6px 10px; }}
  tbody tr:nth-child(even) {{ background: #f6f8fa; }}
  tr:hover {{ background: #fff3cd; }}
  .muted {{ color: #6b7280; font-size: 13px; }}
  .filter {{ background: #e6f4ea; padding: 8px 14px; border-radius: 4px; display: inline-block; margin-bottom: 16px; }}
  .issues {{ display: flex; flex-direction: column; gap: 14px; margin: 16px 0 24px; }}
  .issue {{ border: 1px solid #e1e4e8; border-radius: 8px; overflow: hidden; }}
  .issue-head {{ padding: 12px 16px; color: #fff; font-weight: 600; font-size: 15px; }}
  .issue-sev1 {{ background: #dc2626; }}
  .issue-sev2 {{ background: #ea580c; }}
  .issue-sev3 {{ background: #ca8a04; }}
  .issue-body {{ padding: 12px 16px; background: #fff; }}
  .issue-summary {{ font-size: 14px; line-height: 1.7; margin-bottom: 8px; }}
  .issue-evidence {{ margin: 6px 0 0; padding-left: 20px; }}
  .issue-evidence li {{ font-size: 13px; color: #475569; line-height: 1.6; margin: 3px 0; }}
</style></head><body>
<h1>CI 耗时分析报告</h1>
<div class="meta">
  仓库：<b>{html_lib.escape(repo)}</b> ｜ 工作流：<b>{html_lib.escape(workflow)}</b><br>
  时间范围：{html_lib.escape(date_from)} ~ {html_lib.escape(date_to)} ｜ 生成时间：{datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}
</div>
<div class="filter">过滤标准：conclusion = success 且 耗时 &gt; {min_duration} 分钟（排除纯 lint 快速通过的 run）</div>
{summary_cards}
{issues_html}
<h2>📊 统计数据</h2>
{_stats_table("Job 耗时统计（按总耗时降序，仅成功 job）", job_rows, "全量数据见 Excel job_stats sheet")}
{_stats_table("Step 耗时统计（按总耗时降序，仅成功 step）", step_rows, "全量数据见 Excel step_stats sheet")}
{_stats_table("Run 明细", run_rows, "按创建时间排序")}
<p class="muted">API 调用：{api_calls} 次 ｜ 耗时：{elapsed:.1f}s ｜ 原始明细数据见同目录 Excel 文件</p>
</body></html>"""
    Path(filepath).write_text(doc, encoding="utf-8")
    print(f"✅ HTML 报告: {filepath}", file=sys.stderr)


# ─── 主流程 ────────────────────────────────────────────────────────────

def main() -> int:
    ap = argparse.ArgumentParser(description="CI 耗时分析（目的2）：成功+跑了所有job 的 run 的 job/step 耗时")
    ap.add_argument("--repo", required=True, help="owner/name")
    ap.add_argument("--workflow", required=True, help="workflow 显示名（如 E2E）")
    ap.add_argument("--from", dest="date_from", required=True, help="起始日期 YYYY-MM-DD")
    ap.add_argument("--to", dest="date_to", required=True, help="结束日期 YYYY-MM-DD")
    ap.add_argument("--min-duration", type=float, default=DEFAULT_MIN_DURATION, help=f"最小耗时阈值(分钟)，默认 {DEFAULT_MIN_DURATION}")
    ap.add_argument("--output-dir", default=".", help="输出目录")
    ap.add_argument("--token", default=os.getenv("GITHUB_TOKEN") or os.getenv("GH_TOKEN"))
    ap.add_argument("--concurrency", type=int, default=8)
    ap.add_argument("--step-names", default=str(DEFAULT_STEP_NAMES))
    args = ap.parse_args()
    if not args.token:
        print("需要 --token 或 GITHUB_TOKEN/GH_TOKEN（可用 gh auth token）", file=sys.stderr)
        return 2

    step_map = {}
    if Path(args.step_names).exists():
        step_map = json.load(open(args.step_names))
        print(f"📂 step 分类: {len(step_map)} 条映射", file=sys.stderr)

    client = GitHubClient(args.token)
    wf = fetch_workflow_id(client, args.repo, args.workflow)
    if not wf:
        print(f"未找到名称含 '{args.workflow}' 的 active workflow", file=sys.stderr)
        return 1
    wid, wname = wf
    print(f"workflow: {wname} (id={wid})", file=sys.stderr)

    t0 = time.time()
    # 1. 拉 run 列表（少量分页调用）
    runs = fetch_runs_in_range(client, args.repo, wid, args.date_from, args.date_to)
    print(f"{args.date_from}~{args.date_to} 共 {len(runs)} 个 run（API {client.calls}，{time.time()-t0:.1f}s）", file=sys.stderr)
    if not runs:
        return 1

    # 2. 预过滤：只留成功 + >min_duration（省 job 抓取的 API 调用）
    filtered = filter_runs(runs, args.min_duration)
    concl = defaultdict(int)
    for r in runs:
        concl[r.conclusion or "unknown"] += 1
    print(f"过滤后命中 {len(filtered)}/{len(runs)} 个 run（结论分布: {dict(concl)}）", file=sys.stderr)
    if not filtered:
        print("⚠ 过滤后无 run，可降低 --min-duration 或扩大时间范围", file=sys.stderr)
        return 1

    # 3. 只对命中 run 抓 jobs(含 steps)
    print(f"并发抓取 {len(filtered)} 个 run 的 jobs（concurrency={args.concurrency}）...", file=sys.stderr)
    run_jobs = fetch_jobs_for_runs(client, args.repo, [r.id for r in filtered], args.concurrency)
    # 把 jobs 挂回 run 对象，便于明细输出复用
    for r in filtered:
        r.jobs = run_jobs.get(r.id, [])
    elapsed = time.time() - t0
    njobs = sum(len(v) for v in run_jobs.values())
    nsteps = sum(len(s.steps) for v in run_jobs.values() for s in v)
    print(f"jobs 抓取完成: {njobs} jobs / {nsteps} steps，API {client.calls}，{elapsed:.1f}s", file=sys.stderr)

    # 4. 输出
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    safe_repo = args.repo.replace("/", "_")
    tag = f"{args.date_from}_{args.date_to}"
    xlsx_path = out_dir / f"{safe_repo}-{args.workflow}-duration-raw-{tag}.xlsx"
    html_path = out_dir / f"{safe_repo}-{args.workflow}-duration-report-{tag}.html"

    write_excel(str(xlsx_path), filtered, run_jobs, step_map)
    write_html(str(html_path), args.repo, wname, args.date_from, args.date_to,
               len(runs), filtered, run_jobs, step_map, args.min_duration, client.calls, elapsed)
    print(f"\n📊 完成：HTML 报告 + Excel 原始数据已生成于 {out_dir}", file=sys.stderr)
    return 0


# ─── 自检（ponytail：非平凡逻辑留一个可运行检查）─────────────────────

def _selfcheck():
    """验证过滤逻辑与洞察生成的基本正确性。运行: python3 ci_duration_analysis.py --selfcheck"""
    from datetime import datetime as _dt
    base = _dt(2026, 6, 29, 10, 0, tzinfo=timezone.utc)

    def _run(conc, dur_min, rid=1):
        return WorkflowRunInfo(
            id=rid, name="E2E", workflow_id=1, status="completed", conclusion=conc,
            event="pull_request", html_url="", head_sha="",
            created_at=base, run_started_at=base,
            updated_at=base + timedelta(minutes=dur_min),
        )

    runs = [
        _run("success", 90, 1),     # 命中
        _run("success", 4, 2),      # 不命中：lint 快速通过
        _run("failure", 90, 3),     # 不命中：失败
        _run("cancelled", 90, 4),   # 不命中：取消
    ]
    got = filter_runs(runs, 10.0)
    assert [r.id for r in got] == [1], f"filter_runs 应只留 success+>10min，got {[r.id for r in got]}"

    # 洞察生成在空数据时不崩
    ins = generate_insights([], {}, {})
    assert isinstance(ins, list) and ins

    print("OK: selfcheck 通过（filter_runs + generate_insights 基本逻辑正确）")


if __name__ == "__main__":
    if "--selfcheck" in sys.argv:
        _selfcheck()
        raise SystemExit(0)
    raise SystemExit(main())
